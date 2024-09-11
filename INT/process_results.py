import argparse
from cmath import sqrt
import csv
import os
import sys
import ast

from pprint import pprint
from influxdb import InfluxDBClient
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Define the directory path
result_directory = "results"
analyzer_directory = "analyzer"
final_file = "final_results.xlsx"
current_directory = os.path.dirname(os.path.abspath(__file__)) 

args = None
results = {}

num_switches = 14           #switches ids go from 1 to 14

# Define DB connection parameters
host='localhost'
dbname='int'
# Connect to the InfluxDB client
client = InfluxDBClient(host=host, database=dbname)

# SRv6 logs
log_file = "ECMP-SRv6 rules.log"

algorithms = ["KShort", "ECMP", "ECMP-SRv6"]

headers_lines = ["AVG Out of Order Packets (Nº)", "AVG Packet Loss (Nº)", "AVG Packet Loss (%)", 
                "AVG 1º Packet Delay (nanoseconds)", 
                "AVG Nº of SRv6 rules Created", "AVG Nº of SRv6 rules Removed",
                "AVG Flows Latency (nanoseconds)", "STD Flows Latency (nanoseconds)", 
                "AVG Hop Latency (nanoseconds)", "STD Hop Latency (nanoseconds)",
                "AVG of packets to each switch (%)", 
                "Standard Deviation of packets to each switch (%)", 
                "AVG of processed Bytes to each switch", "Standard Deviation of processed Bytes to each switch", 
                "Variation of the AVG 1º Packet Delay between (No)Emergency Flows (%)",
                "Variation of the AVG Flow Delay between (No)Emergency Flows (%)"]

num_values_to_compare_all_tests = len(headers_lines)

def parse_args():
    global args

    parser = argparse.ArgumentParser(description='process parser')
    parser.add_argument('--f', help='CSV files to be processed',
                        type=str, action="store", required=True, nargs='+')
    
    # 2 arguments with multiple values, the nº of elements must be the same between them and the files
    parser.add_argument('--start', help='Timestamp (RFC3339 format) of when each test started (1 peer file)',
                        type=str, action="store", required=True, nargs='+')
    parser.add_argument('--end', help='Timestamp (RFC3339 format) of when each test ended (1 peer file)',
                        type=str, action="store", required=True, nargs='+')
    
    # 3 Optional argument, that must be used simultaneously
    parser.add_argument('--SRv6_index', help='Indexs of the files that have SRv6 logs associated to them (starting from 0)',
                        type=int, action="store", required=False, nargs='+')
    parser.add_argument('--SRv6_logs', help='Names of the log files with the SRv6 rules from previous argument (must match the order)',
                        type=str, action="store", required=False, nargs='+')
    parser.add_argument('--num_iterations', help='Nº of iterations on every test, help SRv6 AVG calculations)',
                        type=int, action="store", required=False)

    args = parser.parse_args()
    
    # Check if the number of elements start/end is the same
    if len(args.start) != len(args.end):
        parser.error("The number of elements in --start and --end must be the same")
    if len(args.start) != len(args.f):
        parser.error("The number of elements in --start and --end must be the same as the number of files")


    # Check if args pairs SRv6_index, SRv6_logs and num_iterations are used together
    if (args.SRv6_index and not args.SRv6_logs) or (not args.SRv6_index and args.SRv6_logs):
        parser.error("Both --SRv6_index and --SRv6_logs must be used together")
    if not args.num_iterations and args.SRv6_index:
        parser.error("The --num_iterations must be used with --SRv6_index and --SRv6_logs")

    # Check if the number of elements is the same in SRv6_index and SRv6_logs
    if args.SRv6_index and args.SRv6_logs:
        if len(args.SRv6_index) != len(args.SRv6_logs):
            parser.error("The number of elements in --SRv6_index and --SRv6_logs must be the same")

    # If index of SRv6 is given, check if it is valid
    if args.SRv6_index:
        for index in args.SRv6_index:
            if index < 0 or index >= len(args.f):
                parser.error("The SRv6_index: "+ str(index) +" is invalid. It must be between 0 and the number of files-1")

def apply_query(query):
    global client
    try:
        # Execute the query
        result = client.query(query)
    except Exception as error:
        # handle the exception
        print("An exception occurred:", error)

    

    return result

def extract_Is_values(line, iteration, flow, Is):
    line = line + [Is]
    # Is values, all in the current line
    for key, Is_value in results[iteration][flow][Is].items():
        if key == "extra":
            for extra in Is_value:
                line = line + list(extra.values())
        else:
            line = line + [Is_value]

    return line

def adjust_columns_width():
    global final_file
    #open the workbook
    dir_path = os.path.join(current_directory, result_directory)
    file_path = os.path.join(dir_path, final_file)
    workbook = load_workbook(file_path)

    # Adjust column widths to fit the text
    for sheetname in workbook.sheetnames:
        print(f"Adjusting columns width for sheet {sheetname}")
        sheet = workbook[sheetname]
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
    
    # Save the workbook
    workbook.save(file_path)

def get_pkt_size_dscp(flow):
    #reads the INT DB and sets the pkt size and DSCP collumns

    query = f"""
        SELECT dscp, size
        FROM flow_stats
        WHERE   src_ip = '{flow[0]}'
        AND     dst_ip = '{flow[1]}'
        AND     flow_label = '{flow[2]}'
        ORDER BY time DESC
        LIMIT 1
    """
    
    r = apply_query(query)
    
    dscp = r.raw["series"][0]["values"][0][1]
    size = r.raw["series"][0]["values"][0][2]

    #print(dscp)
    #print(size)

    return dscp, size

def read_raw_results(row):
    global results
    iteration = row[0]
    flow = (row[1], row[2], int(row[3]))
    Is = row[4]
    number_of_packets = int(row[5])
    first_packet_time = row[6]
    values_end_points = {}
    values_end_points["num_pkt"] = number_of_packets
    values_end_points["time"] = float(first_packet_time)


    if Is == "receiver":
        number_out_of_order_packets = int(row[7])
        out_of_order_packets = row[8]

        extra1 = {"num_out_of_order_pkt": number_out_of_order_packets}
        extra2 = {"out_of_order_pkt": out_of_order_packets}

        values_end_points["extra"] = [extra1, extra2]

    # Check if the iteration is already in the results dictionary
    if iteration not in results:
        dscp, pkt_size = get_pkt_size_dscp(flow)            #Get the flows info
        values_flow = {Is: values_end_points, "DSCP": dscp, "Packet Size": pkt_size}
        values_iteration = {flow: values_flow}
        results[iteration] = values_iteration
    else:
        # Check if the flow is already in the results dictionary
        if flow not in results[iteration]:
            dscp, pkt_size = get_pkt_size_dscp(flow)            #Get the flows info
            values_flow = {Is: values_end_points, "DSCP": dscp, "Packet Size": pkt_size}
            results[iteration][flow] = values_flow
        else:
            # Add currect Is to the flow
            results[iteration][flow][Is] = values_end_points

def read_csv_files(filename):
    first_row = True

    res_path = os.path.join(current_directory, result_directory) 
    file_path = os.path.join(res_path, filename)

    # Check if the directory exists
    if not os.path.isdir(res_path):
        sys.exit(1)

    #check file exists
    if not os.path.isfile(file_path):
        print(f"File {filename} not found in {file_path}")
        sys.exit(1)
        
    print(f"Reading files: {filename}")

    try:
        with open(file_path, 'r', newline='') as csvfile:
            reader = csv.reader(csvfile)
            for row in reader:
                #Skip the header, first row
                if first_row:
                    first_row = False
                    continue
                read_raw_results(row)
    except Exception as e:
        print(f"An error occurred while reading {filename}: {e}")
    #print("Done reading file")
    #pprint(results)

def read_SRv6_line(line):
    global results
    #print(f"Reading SRv6 line: {line}")

    new_content = {}
    part1 = line.split(" - ")
    part2 = part1[2].split(" => ")
    operation = part2[0]
    part3 = part2[1].split(": ")
    iteration = part1[0]
    timestamp = part1[1]
    operation = part2[0]
    responsible_switch = int(part3[0])
    rule_elemets_str = line.split("{")[1]
    rule_elemets = ast.literal_eval("{"+rule_elemets_str)

    # Convert numebrs in the dictionary to integers
    rule_elemets["deviceID"] = int(rule_elemets["deviceID"])
    rule_elemets["flow_label"] = int(rule_elemets["flow_label"])
    rule_elemets["src_mask"] = int(rule_elemets["src_mask"])
    rule_elemets["dst_mask"] = int(rule_elemets["dst_mask"])
    rule_elemets["flow_label_mask"] = int(rule_elemets["flow_label_mask"])

    # Extract some values from the dictionary
    src_mask = rule_elemets["src_mask"]
    dst_mask = rule_elemets["dst_mask"]
    flow_label_mask = rule_elemets["flow_label_mask"]

    if src_mask != 128 or dst_mask != 128 or flow_label_mask != 255:
        print(f"Error in the masks in the rule: {rule_elemets}, we only work with flow specific rules")
        sys.exit(1)


    #---------------Create the new operation for the dictionary
    new_content["responsible_switch"] = responsible_switch    # Which switch the load induced the operation
    new_content["operation"] = operation
    new_content["timestamp"] = timestamp
    new_content["rule"] = rule_elemets

    #print("--------------------------------------------------------")
    #print(f"Iteration: {iteration}, Timestamp: {timestamp}, Operation: {operation}, Responsible switch: {responsible_switch}")
    #pprint(rule_elemets)
    #pprint(results)
    #pprint(new_content)


    # Add the results to the dictionary
    # At this point the iteration is already in the dictionary
    results_ite = results[iteration]
    
    # Check if the operation if "SRv6_Operations" is already in the current iteration
    if "SRv6_Operations" not in results_ite:
        results_ite["SRv6_Operations"] = [new_content]   # as a list since there can be multiple operations
    else:
        # Add the operation to the dictionary
        results_ite["SRv6_Operations"].append(new_content)

    #pprint(results)

def read_SRv6_log(file_index):
    analyzer_logs_dir = os.path.join(current_directory, analyzer_directory)
    
    #Get the position of the given file_index in args.SRv6_index
    log_index = args.SRv6_index.index(file_index)

    # Read the SRv6 logs
    log_file = args.SRv6_logs[log_index]
    log_file_path = os.path.join(analyzer_logs_dir, log_file)

    print(f"Reading SRv6 logs from file: {log_file}")

    # Read the SRv6 logs
    with open(log_file_path, 'r') as file:
        for line in file:
            read_SRv6_line(line)

def check_files_exist():
    # Check if the directory/files exist
    full_res_path = os.path.join(current_directory, result_directory) 

    if not os.path.isdir(full_res_path):  # Correct condition to check if the directory does not exist
        print(f"Directory {result_directory} does not exist.")
        sys.exit(1)

    for filename in args.f:
        file_path = os.path.join(full_res_path, filename)
        # Check if the file exists
        if not os.path.isfile(file_path):
            print(f"File {filename} not found in {file_path}")
            sys.exit(1)


    #---------------Check if the SRv6 logs exist

    full_analy_path = os.path.join(current_directory, analyzer_directory) 

    # Check if the directory exists
    if not os.path.isdir(full_analy_path):
        print(f"Directory {analyzer_directory} does not exist.")
        sys.exit(1)
    
    # Check if the log files exist
    if args.SRv6_logs is not None:
        for log_file in args.SRv6_logs:
            log_file_path = os.path.join(full_analy_path, log_file)
            #print(f"Checking SRv6 log file: {log_file_path}")

            if not os.path.isfile(log_file_path):
                print(f"File {log_file} not found in {log_file_path}")
                sys.exit(1)

def export_SRv6_rules(sheet, iteration):
    sheet.append([""])

    # Add the "SRv6 Operations" header
    header = ["SRv6 Operations"]
    new_next_row = sheet.max_row + 1
    for col_num, value in enumerate(header, 1):
        cell = sheet.cell(row=new_next_row, column=col_num, value=value)
        cell.font = Font(bold=True)
    
    # Add the detailed headers for SRv6 operations
    header = ["Timestamp", "Operation", "Responsible Switch", "Source", "Destination", "Flow Label"]
    new_next_row = sheet.max_row + 1
    for col_num, value in enumerate(header, 1):
        cell = sheet.cell(row=new_next_row, column=col_num, value=value)
        cell.font = Font(bold=True)

    # Add the SRv6 operations
    for operation in results[iteration]["SRv6_Operations"]:
        new_next_row = sheet.max_row + 1
        rule_src_IP = operation["rule"]["srcIP"]
        rule_dst_IP = operation["rule"]["dstIP"]
        rule_flow_label = operation["rule"]["flow_label"]
        line = [operation["timestamp"], operation["operation"], operation["responsible_switch"], rule_src_IP, rule_dst_IP, rule_flow_label]
        sheet.append(line)

    sheet.append([""])

def export_results(OG_file):
    global results
    # Get the sheet name from filename before (_)
    sheet_name = OG_file.split("_")[0]

    dir_path = os.path.join(current_directory, result_directory)

    # Ensure the result directory exists
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    
    # Check if the file exists
    file_path = os.path.join(dir_path, final_file)
    if os.path.exists(file_path):
        # Load the existing workbook
        workbook = load_workbook(file_path)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(title=sheet_name)
    else:
        # Create a new workbook and sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
    
    # Write the header
    header = ["Flow src", "Flow dst", "Flow Label", "DSCP", "Packet Size (Bytes)", "Is", "Nº of packets", "1º Packet Timestamp(seconds)", "Nº of out of order packets", "Out of order packets"]
    for col_num, value in enumerate(header, 1):
        cell = sheet.cell(row=1, column=col_num, value=value)
        cell.font = Font(bold=True)

    # Write results, iteration by iteration
    for iteration in results:
        
        # Write key
        sheet.append([f""])
        sheet.append([None])  # This adds an empty row
        cell = sheet.cell(row=sheet.max_row, column=1)  # Get the last row and column 1
        cell.value = f"Iteration - {iteration}"
        cell.font = Font(bold=True)
        # Flow by flow
        for flow in results[iteration]:
            if flow == "SRv6_Operations":       #It is not a flow,
                continue
            keys = list(results[iteration][flow].keys())

            DSCP = results[iteration][flow]["DSCP"]
            pkt_size =  results[iteration][flow]["Packet Size"]

            # Check if both types of Is are keys in the dictionary
            if "sender" not in keys:
                print(f"Sender not found in iteration {iteration} flow {flow}")
                sys.exit(1)
            if "receiver" not in keys:
                print(f"Receiver not found in iteration {iteration} flow {flow}")
                sys.exit(1)

            # Is by Is, sender must be the 1º
            OG_line = list(flow) + [DSCP, pkt_size]
            line = extract_Is_values(OG_line, iteration, flow, "sender")
            sheet.append(line)

            OG_line = list(flow) + [DSCP, pkt_size]
            line = extract_Is_values(OG_line, iteration, flow, "receiver")
            sheet.append(line)

        # Write the SRv6 operations of this Iteration if they exist
        if "SRv6_Operations" in results[iteration]:
            export_SRv6_rules(sheet, iteration)

    # Save the workbook
    workbook.save(file_path)

def set_pkt_loss():
    # Configure each sheet
    dir_path = os.path.join(current_directory, result_directory)
    file_path = os.path.join(dir_path, final_file)
    workbook = load_workbook(file_path)

    # Set formula for each sheet
    for sheet in workbook.sheetnames:
        sheet = workbook[sheet]

        #Set new headers
        sheet['L1'] = "Packet Loss"
        sheet['M1'] = "Packet Loss (%)"

        sheet['L1'].font = Font(bold=True)
        sheet['M1'].font = Font(bold=True)

        no_formula_section = False

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
            
            # Skip the SRv6 Operations section
            if row[0].value == "SRv6 Operations":
                no_formula_section = True
            elif row[0].value and row[0].value.startswith("Iteration -"):
                no_formula_section = False
            if no_formula_section:
                continue

            #if cell from collumn A does not contain an IPv6 address, skip
            if row[0].value is None or ":" not in row[0].value:
                skip = True
                continue
            if skip:            #not in the right line of the pair
                skip = False
                continue
            #print(row)
            
            # Set the formula, pkt loss, -1 is sender, 0 is receiver
            sheet[f'L{row[0].row}'] = f'=G{row[0].row-1}-G{row[0].row}'     
            sheet[f'M{row[0].row}'] = f'=ROUND((L{row[0].row}/G{row[0].row-1})*100, 3)'

            skip = True

    # Save the workbook
    workbook.save(file_path)

def set_fist_pkt_delay():
    # Configure each sheet
    dir_path = os.path.join(current_directory, result_directory)
    file_path = os.path.join(dir_path, final_file)
    workbook = load_workbook(file_path)
    
    # Set formula for each sheet
    for sheet in workbook.sheetnames:
        sheet = workbook[sheet]

        #Set new headers as bold text
        sheet['N1'] = "1º Packet Delay (nanoseconds)"
        sheet['N1'].font = Font(bold=True)

        no_formula_section = False
        
        # Set collumn L to contain a formula to be the subtraction of values of collum F of the current pair of lines
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):

            # Skip the SRv6 Operations section
            if row[0].value == "SRv6 Operations":
                no_formula_section = True
            elif row[0].value and row[0].value.startswith("Iteration -"):
                no_formula_section = False
            if no_formula_section:
                continue

            #if cell from collumn A does not contain an IPv6 address, skip
            if row[0].value is None or ":" not in row[0].value:
                skip = True
                continue
            if skip:            #not in the right line of the pair
                skip = False
                continue
            #print(row)
            
            # Set the formula, pkt loss, -1 is sender, 0 is receiver
            # The values are 2 Timestamp (seconds-Unix Epoch)
            # subtraction give seconds, we convert to nanoseconds
            sheet[f'N{row[0].row}'] = f'=ROUND((H{row[0].row}-H{row[0].row-1})*10^9, 3)'     

            skip = True

    # Save the workbook
    workbook.save(file_path)

def set_caculations():
    # Configure each sheet
    dir_path = os.path.join(current_directory, result_directory)
    file_path = os.path.join(dir_path, final_file)
    workbook = load_workbook(file_path)

    # Set formula for each sheet
    for sheet in workbook.sheetnames:
        sheet = workbook[sheet]

        #Pass the last line with data, and leave 2 empty lines
        last_line = sheet.max_row + 4

        #Set new headers
        sheet[f'A{last_line}'] = "Calculations"
        sheet[f'A{last_line + 1}'] = "AVG Out of Order Packets (Nº)"
        sheet[f'A{last_line + 2}'] = "AVG Packet Loss (Nº)"
        sheet[f'A{last_line + 3}'] = "AVG Packet Loss (%)"
        sheet[f'A{last_line + 4}'] = "AVG 1º Packet Delay (nanoseconds)"
        sheet[f'A{last_line + 5}'] = "AVG Nº of SRv6 rules Created"
        sheet[f'A{last_line + 6}'] = "AVG Nº of SRv6 rules Removed"
        sheet[f'B{last_line}'] = "Values"

        sheet[f'A{last_line}'].font = Font(bold=True)
        sheet[f'A{last_line + 1}'].font = Font(bold=True)
        sheet[f'A{last_line + 2}'].font = Font(bold=True)
        sheet[f'A{last_line + 3}'].font = Font(bold=True)
        sheet[f'A{last_line + 4}'].font = Font(bold=True)
        sheet[f'A{last_line + 5}'].font = Font(bold=True)
        sheet[f'A{last_line + 6}'].font = Font(bold=True)
        sheet[f'B{last_line}'].font = Font(bold=True)

        # on the next line for each column, set the average of the column, ignore empty cells
        sheet[f'B{last_line + 1}'] = f'=ROUND(AVERAGEIF(I:I, "<>", I:I), 3)'
        sheet[f'B{last_line + 2}'] = f'=ROUND(AVERAGEIF(L:L, "<>", L:L), 3)'
        sheet[f'B{last_line + 3}'] = f'=ROUND(AVERAGEIF(M:M, "<>", M:M), 3)'
        sheet[f'B{last_line + 4}'] = f'=ROUND(AVERAGEIF(N:N, "<>", N:N), 3)'
        sheet[f'B{last_line + 5}'] = f'=COUNTIF(B1:B{last_line}, "Created SRv6 rule") / {args.num_iterations}'
        sheet[f'B{last_line + 6}'] = f'=COUNTIF(B1:B{last_line}, "Removed SRv6 rule") / {args.num_iterations}'



    # Save the workbook
    workbook.save(file_path)

def get_byte_sum(start, end):
    # Initialize the result dictionary to store total byte counts per switch ID
    sum = {}
    switch_ids = []

    #--------------Query to get unique switch_id values if it is a tag
    # This approche did not take into account the existing but unused switches
    #query = f'''
    #        SHOW TAG VALUES 
    #        FROM "switch_stats"  
    #        WITH KEY = "switch_id"
    #        WHERE time >= '{start}' AND time <= '{end}' 
    #    '''
    #tmp = apply_query(query)  

    # Extract the switch_id values into a list
    #for row in tmp.raw["series"][0]["values"]:
    #    switch_ids.append(int(row[1]))

    switch_ids = list(range(1, num_switches + 1))

    #pprint(switch_ids)

    # Loop through each unique switch ID
    for switch_id in switch_ids:
        # Formulate the query to get the sum of bytes for the given switch ID and time range
        query = f"""
            SELECT SUM("size") AS total_count
            FROM flow_stats 
            WHERE time >= '{start}' AND time <= '{end}' 
            AND path =~ /(^|-)({switch_id})(-|$|\b)/
        """

        result = apply_query(query)  # Assume this returns a dictionary with 'total_count'
        
        # Add the result to the sum dictionary under the switch_id key
        #print(f"Switch ID: {switch_id}")
        #print(result.raw)
        sum[switch_id] = {}
        if not result.raw["series"]:
            sum[switch_id]["Byte Sums"] = 0
        else:
            sum[switch_id]["Byte Sums"] = result.raw["series"][0]["values"][0][1]
    #pprint(sum)

    return sum

def calculate_percentages(start, end, switch_data):
    # Get the total count of packets
    query = f"""
        SELECT COUNT("latency") AS total_count
        FROM flow_stats
        WHERE time >= '{start}' AND time <= '{end}'
    """
    result = apply_query(query)
    total_count = result.raw["series"][0]["values"][0][1]  # Extract total_count from the result

    # Get the count of packets that went to each switch
    query = f"""
        SELECT COUNT("latency") AS switch_count
        FROM switch_stats
        WHERE time >= '{start}' AND time <= '{end}'
        GROUP BY switch_id
    """
    result = apply_query(query)
    
    # Calculate the percentage of packets that went to each switch
    # initialize to all switches as 0, so unused switches are takrn into account too
    for switch_id in range(1, num_switches + 1):
        switch_data[switch_id]["Percentage Pkt"] = 0

    for row in result.raw["series"]:
        #tuple pair: id, count
        switch_id = int(row["tags"]["switch_id"])
        switch_count = int(row["values"][0][1])
        switch_data[switch_id]["Percentage Pkt"] = round((switch_count / total_count) * 100, 3)

    #pprint(switch_data)
    return switch_data

def get_mean_standard_deviation(switch_data):
    #pprint(switch_data)
    sum_percentage = 0
    sum_byte = 0
    count = 0

    # Get the mean of the (Byte Sums) and (Percentage Pkt) for all switches in switch_data
    for switch_id in switch_data:
        sum_percentage += switch_data[switch_id]["Percentage Pkt"]
        sum_byte += switch_data[switch_id]["Byte Sums"]
        count += 1

    percentage_mean = sum_percentage / count
    byte_mean = sum_byte / count
    

    # Get the Standard Deviation of the (Byte Sums) and (Percentage Pkt) for all switches in switch_data
    sum_squared_diff_percentage = 0
    sum_squared_diff_byte = 0
    
    for switch_id in switch_data:
        current_percentage = switch_data[switch_id]["Percentage Pkt"]
        current_byte = switch_data[switch_id]["Byte Sums"]
        sum_squared_diff_percentage += (current_percentage - percentage_mean) ** 2
        sum_squared_diff_byte += (current_byte - byte_mean) ** 2
    
    percentage_std_dev = sqrt(sum_squared_diff_percentage / count).real
    byte_std_dev = sqrt(sum_squared_diff_byte / count).real

    #print("percentage_std_dev: ", percentage_std_dev)
    #print("byte_std_dev: ", percentage_std_dev)
    
    switch_data["Percentage Mean"] = round(percentage_mean, 3)
    switch_data["Byte Mean"] = round(byte_mean, 3)
    switch_data["Percentage Standard Deviation"] = round(percentage_std_dev, 3)
    switch_data["Byte Standard Deviation"] = round(byte_std_dev, 3)

    return switch_data

def set_test_case_headers(sheet, test_case, start_line):
    global headers_lines

    # Set test case name in bold test
    title = f"{test_case}"
    sheet[f'A{start_line}'] = title
    sheet[f'A{start_line}'].font = Font(bold=True)

    # Set the collumn names
    sheet[f'B{start_line}'] = "KShort"
    sheet[f'C{start_line}'] = "ECMP"
    sheet[f'D{start_line}'] = "ECMP+SRv6"
    sheet[f'E{start_line}'] = "Variation1 (%)"
    sheet[f'F{start_line}'] = "Variation2 (%)"
    sheet[f'G{start_line}'] = "Variation3 (%)"

    # Set collumn names in bold text
    sheet[f'B{start_line}'].font = Font(bold=True)
    sheet[f'C{start_line}'].font = Font(bold=True)
    sheet[f'D{start_line}'].font = Font(bold=True)
    sheet[f'E{start_line}'].font = Font(bold=True)
    sheet[f'F{start_line}'].font = Font(bold=True)
    sheet[f'G{start_line}'].font = Font(bold=True)

    # Set the lines names
    sheet[f'A{start_line + 1}'] = headers_lines[0]
    sheet[f'A{start_line + 2}'] = headers_lines[1]
    sheet[f'A{start_line + 3}'] = headers_lines[2]
    sheet[f'A{start_line + 4}'] = headers_lines[3]
    sheet[f'A{start_line + 5}'] = headers_lines[4]
    sheet[f'A{start_line + 6}'] = headers_lines[5]
    sheet[f'A{start_line + 7}'] = headers_lines[6]
    sheet[f'A{start_line + 8}'] = headers_lines[7]
    sheet[f'A{start_line + 9}'] = headers_lines[8]
    sheet[f'A{start_line + 10}'] = headers_lines[9]
    sheet[f'A{start_line + 11}'] = headers_lines[10]
    sheet[f'A{start_line + 12}'] = headers_lines[11]
    sheet[f'A{start_line + 13}'] = headers_lines[12]
    sheet[f'A{start_line + 14}'] = headers_lines[13]
    sheet[f'A{start_line + 15}'] = headers_lines[14]
    sheet[f'A{start_line + 16}'] = headers_lines[15]


    # Set lines names in bold text
    sheet[f'A{start_line + 1}'].font = Font(bold=True)
    sheet[f'A{start_line + 2}'].font = Font(bold=True)
    sheet[f'A{start_line + 3}'].font = Font(bold=True)
    sheet[f'A{start_line + 4}'].font = Font(bold=True)
    sheet[f'A{start_line + 5}'].font = Font(bold=True)
    sheet[f'A{start_line + 6}'].font = Font(bold=True)
    sheet[f'A{start_line + 7}'].font = Font(bold=True)
    sheet[f'A{start_line + 8}'].font = Font(bold=True)
    sheet[f'A{start_line + 9}'].font = Font(bold=True)
    sheet[f'A{start_line + 10}'].font = Font(bold=True)
    sheet[f'A{start_line + 11}'].font = Font(bold=True)
    sheet[f'A{start_line + 12}'].font = Font(bold=True)
    sheet[f'A{start_line + 13}'].font = Font(bold=True)
    sheet[f'A{start_line + 14}'].font = Font(bold=True)
    sheet[f'A{start_line + 15}'].font = Font(bold=True)
    sheet[f'A{start_line + 16}'].font = Font(bold=True)

def set_comparasion_formulas(sheet, start_line):
    # Set the formulas to compare the results between the test cases
    for i in range(1, num_values_to_compare_all_tests + 1):
        sheet[f'E{start_line + i}'] = f'=IFERROR(ROUND((C{start_line + i} - B{start_line + i}) / ABS(B{start_line + i}) * 100, 3), 0)'
        sheet[f'F{start_line + i}'] = f'=IFERROR(ROUND((D{start_line + i} - B{start_line + i}) / ABS(B{start_line + i}) * 100, 3), 0)'
        sheet[f'G{start_line + i}'] = f'=IFERROR(ROUND((D{start_line + i} - C{start_line + i}) / ABS(C{start_line + i}) * 100, 3), 0)'

def get_line_column_to_copy_from(sheet_to_copy_from_name, variable_number):
    global headers_lines
    line =None
    col = None

    file_path = os.path.join(current_directory, result_directory, final_file)
    workbook = load_workbook(file_path)
    sheet_to_copy_from = workbook[sheet_to_copy_from_name]

    variable_name = headers_lines[variable_number]

    pass_1_occurance = True          #there are 2 Lines on collumn A that have the same name
    if variable_number == 14:
        pass_1_occurance = False 

    # sheet_to_copy_from, get the line of the cell that contains the variable_name on collumn A and the collumn after it
    for row in sheet_to_copy_from.iter_rows(min_row=1, max_row=sheet_to_copy_from.max_row, min_col=1, max_col=1):
        
        if pass_1_occurance == False and row[0].value == "AVG 1º Packet Delay (nanoseconds)":
            pass_1_occurance = True
            continue
        
        if variable_number <= 9:
            if row[0].value == variable_name:
                # Get the next collumn letter of the cell that contains the variable_name
                line = row[0].row
                col = get_column_letter(row[0].column + 1)
                break
        elif variable_number ==10 or variable_number == 12:
            if row[0].value == "Mean":
                line = row[0].row
                if variable_number == 10:
                    col = get_column_letter(row[0].column + 1)
                else:
                    col = get_column_letter(row[0].column + 2)
                break
        elif variable_number == 11 or variable_number == 13:
            if row[0].value == "Standard Deviation":
                line = row[0].row
                if variable_number == 11:
                    col = get_column_letter(row[0].column + 1)
                else:
                    col = get_column_letter(row[0].column + 2)
                break
        elif variable_number == 14:
            if row[0].value == "AVG 1º Packet Delay (nanoseconds)":
                line = row[0].row
                col = get_column_letter(row[0].column + 3)
                break
        elif variable_number == 15:
            if row[0].value == "AVG Flow Delay (nanoseconds)":
                line = row[0].row
                col = get_column_letter(row[0].column + 3)
                break

    return line, col

def set_copied_values(sheet, test_case, start_line):
    global algorithms
    
    # Cycle through the variables to compare (lines)
    for variable_number in range(num_values_to_compare_all_tests):
        
        # Cycle through the algorithms to copy the values (columns)
        for i in range(len(algorithms)):
            #--------------Collumn C is the second algorithm, ECMP
            sheet_to_copy_from_name = f"{test_case}-{algorithms[i]}"
            line, column = get_line_column_to_copy_from(sheet_to_copy_from_name, variable_number)

            if line is None or column is None:
                print(f"Error getting line and column to copy from, sheet_to_copy_from: {sheet_to_copy_from_name}, variable number: {variable_number}")
                continue

            cell_reference = f"{column}{line}"
            formula = f"='{sheet_to_copy_from_name}'!{cell_reference}"
            sheet[f'{get_column_letter(2 + i)}{start_line + 1 + variable_number}'] = formula


def write_INT_results(file_path, workbook, sheet, AVG_flows_latency, STD_flows_latency, AVG_hop_latency, STD_hop_latency, switch_data):
    # Write the results in the sheet
    last_line = sheet.max_row + 1

    # Set new headers
    sheet[f'A{last_line + 0}'] = "AVG Flows Latency (nanoseconds)"
    sheet[f'A{last_line + 1}'] = "STD Flows Latency (nanoseconds)"
    sheet[f'A{last_line + 2}'] = "AVG Hop Latency (nanoseconds)"
    sheet[f'A{last_line + 3}'] = "STD Hop Latency (nanoseconds)"

    sheet[f'A{last_line + 0}'].font = Font(bold=True)
    sheet[f'A{last_line + 1}'].font = Font(bold=True)
    sheet[f'A{last_line + 2}'].font = Font(bold=True)
    sheet[f'A{last_line + 3}'].font = Font(bold=True)

    sheet[f'B{last_line + 0}'] = AVG_flows_latency
    sheet[f'B{last_line + 1}'] = STD_flows_latency
    sheet[f'B{last_line + 2}'] = AVG_hop_latency
    sheet[f'B{last_line + 3}'] = STD_hop_latency


    sheet[f'A{last_line + 5}'] = "Switch ID"
    sheet[f'B{last_line + 5}'] = "% of packets to each switch"
    sheet[f'C{last_line + 5}'] = "Total Sum of Processed Bytes"

    sheet[f'A{last_line + 5}'].font = Font(bold=True)
    sheet[f'B{last_line + 5}'].font = Font(bold=True)
    sheet[f'C{last_line + 5}'].font = Font(bold=True)


    # Write percentages and total bytes processed, cycle through keys that are numbers
    #pprint(switch_data)
    #print("----------------------------------------")
    for i, key in enumerate(switch_data.keys()):
        if isinstance(key, int):                #skip sets that are non-switch_id
            #print(f"Key: {key} is an integer")
            #print(f"i: {i}, Switch ID: {key},  Values: {switch_data[key]}")
            sheet[f'A{last_line + 6 + i}'] = key
            
            #percentage of total packets that went to each switch
            sheet[f'B{last_line + 6 + i}'] = switch_data[key]["Percentage Pkt"]
            
            #Sum of processed bytes
            sheet[f'C{last_line + 6 + i}'] = switch_data[key]["Byte Sums"]

    # Write the mean and standard deviation of the percentages and bytes
    
    sheet[f'A{last_line + num_switches + 5 + 1}'] = "Mean"
    sheet[f'A{last_line + num_switches + 5 + 2}'] = "Standard Deviation"
    sheet[f'A{last_line + num_switches + 5 + 1}'].font = Font(bold=True)
    sheet[f'A{last_line + num_switches + 5 + 2}'].font = Font(bold=True)

    sheet[f'B{last_line + num_switches + 5 + 1}'] = switch_data["Percentage Mean"]
    sheet[f'B{last_line + num_switches + 5 + 2}'] = switch_data["Percentage Standard Deviation"]
    sheet[f'C{last_line + num_switches + 5 + 1}'] = switch_data["Byte Mean"]
    sheet[f'C{last_line + num_switches + 5 + 2}'] = switch_data["Byte Standard Deviation"]
    

    # Save the workbook
    workbook.save(file_path)

def set_INT_results():
    # For each sheet and respectice file, see the time interval given, get the values from the DB, and set the values in the sheet
        
    # Configure each sheet
    dir_path = os.path.join(current_directory, result_directory)
    file_path = os.path.join(dir_path, final_file)
    workbook = load_workbook(file_path)

    # Get nº each sheet
    for i, sheet in enumerate(workbook.sheetnames):
        print(f"Processing sheet {sheet}, index {i}")
        sheet = workbook[sheet]

        # Get the start and end times
        start = args.start[i]
        end = args.end[i]

        ############################################ Get the results from the DB
        # We need AVG Latency of ALL flows combined (NOT distinguishing between flows)
        # Query to get the 95th percentile latency value, to exclude outliers
        percentile_query = f"""
            SELECT PERCENTILE("latency", 95) AS p_latency
            FROM flow_stats
            WHERE time >= '{start}' AND time <= '{end}'
        """

        percentile_result = apply_query(percentile_query)
        p_latency = list(percentile_result.get_points())[0]['p_latency']   #nanoseconds

        query = f"""
                    SELECT MEAN("latency"), STDDEV("latency")
                    FROM  flow_stats
                    WHERE time >= '{start}' AND time <= '{end}' AND "latency" <= {p_latency}
                """
        result = apply_query(query)
        AVG_flows_latency = round(result.raw["series"][0]["values"][0][1], 3)         #nanoseconds
        STD_flows_latency = round(result.raw["series"][0]["values"][0][2], 3)

        ###########################################
        # We need AVG Latency for processing of ALL packets (NOT distinguishing between switches/flows) 
        # Query to get the 95th percentile latency value, to exclude outliers
        percentile_query = f"""
            SELECT PERCENTILE("latency", 95) AS p_latency
            FROM switch_stats
            WHERE time >= '{start}' AND time <= '{end}'
        """
        
        query = f"""
                    SELECT MEAN("latency"), STDDEV("latency")
                    FROM  switch_stats
                    WHERE time >= '{start}' AND time <= '{end}' AND "latency" <= {p_latency}
                """
        result = apply_query(query)
        AVG_hop_latency = round(result.raw["series"][0]["values"][0][1], 3)         #nanoseconds
        STD_hop_latency = round(result.raw["series"][0]["values"][0][2], 3)         

        # % of packets that went to each individual switch (switch_id)
        switch_data = get_byte_sum(start, end)
        switch_data = calculate_percentages(start, end, switch_data)
        switch_data = get_mean_standard_deviation(switch_data)

        #pprint("AVG_flows_latency: ", AVG_flows_latency)
        #pprint("AVG_hop_latency: ", AVG_hop_latency)
        #pprint("switch_data: ", switch_data)

        write_INT_results(file_path, workbook, sheet, AVG_flows_latency, STD_flows_latency, AVG_hop_latency, STD_hop_latency, switch_data)

def get_flow_delays(start, end):
    # Get the average delay of emergency and non-emergency flows
    query = f"""
        SELECT MEAN("latency") 
        FROM  flow_stats WHERE time >= '{start}' 
        AND time <= '{end}' 
        AND dscp = 46
    """
    result = apply_query(query)
    if not result.raw["series"]:
        avg_emergency_flows_delay = "none"
    else:
        avg_emergency_flows_delay = round(result.raw["series"][0]["values"][0][1], 3)         #nanoseconds

    query = f"""
        SELECT MEAN("latency")
        FROM  flow_stats
        WHERE time >= '{start}' AND time <= '{end}'
        AND dscp != 46
    """

    result = apply_query(query)
    if not result.raw["series"]:
        avg_non_emergency_flows_delay = "none"
    else:
        avg_non_emergency_flows_delay = round(result.raw["series"][0]["values"][0][1], 3)         #nanoseconds

    return avg_emergency_flows_delay, avg_non_emergency_flows_delay 

def set_Emergency_calculation():
    # Configure each sheet
    dir_path = os.path.join(current_directory, result_directory)
    file_path = os.path.join(dir_path, final_file)
    workbook = load_workbook(file_path)

    for i, sheet in enumerate(workbook.sheetnames):
        sheet = workbook[sheet]

        # Set new headers
        max_line = sheet.max_row
        sheet[f'A{max_line + 2}'] = "Flows Types"
        sheet[f'B{max_line + 2}'] = "Non-Emergency Flows"
        sheet[f'C{max_line + 2}'] = "Emergency Flows"
        sheet[f'D{max_line + 2}'] = "Variation (%)"
        
        sheet[f'A{max_line + 3}'] = "AVG 1º Packet Delay (nanoseconds)"
        sheet[f'A{max_line + 4}'] = "AVG Flow Delay (nanoseconds)"

        sheet[f'A{max_line + 2}'].font = Font(bold=True)
        sheet[f'B{max_line + 2}'].font = Font(bold=True)
        sheet[f'C{max_line + 2}'].font = Font(bold=True)
        sheet[f'D{max_line + 2}'].font = Font(bold=True)
        sheet[f'A{max_line + 3}'].font = Font(bold=True)
        sheet[f'A{max_line + 4}'].font = Font(bold=True)

        #print(f"Processing sheet {sheet}")
        #print(f"arg.f: {args.f}")
        #get the index of the args.f which the name starts with the current sheet name
        #print(f"Index: {i}")
        start = args.start[i]
        end = args.end[i]

        avg_emergency_flows_delay, avg_non_emergency_flows_delay = get_flow_delays(start, end)

        # Define the row range to consider
        row_range = max_line - 1  # Rows before the max line

        # Set the formula for the Non-Emergency Flows
        sheet[f'B{max_line + 3}'] = f'=IF(SUMIF(D1:D{row_range}, "<>46", N1:N{row_range}) = 0, "none", SUMIF(D1:D{row_range}, "<>46", N1:N{row_range}))'
        sheet[f'B{max_line + 4}'] = avg_non_emergency_flows_delay

        # Set the formula for the Emergency Flows
        sheet[f'C{max_line + 3}'] = f'=IF(SUMIF(D1:D{row_range}, 46, N1:N{row_range}) = 0, "none", SUMIF(D1:D{row_range}, 46, N1:N{row_range}))'
        sheet[f'C{max_line + 4}'] = avg_emergency_flows_delay

        #Set comparasion formulas, for the AVG 1º Packet Delay and AVG Flow Delay in percentage
        sheet[f'D{max_line + 3}'] = f'=IFERROR(ROUND((C{max_line + 3} - B{max_line + 3})/ABS(B{max_line + 3}) * 100, 3), "none")'
        sheet[f'D{max_line + 4}'] = f'=IFERROR(ROUND((C{max_line + 4} - B{max_line + 4})/ABS(B{max_line + 4}) * 100, 3), "none")'


    workbook.save(file_path)

def set_Comparison_sheet():
    print("Setting the Comparison sheet")
    test_cases = ["MEDIUM", "HIGH", "HIGH+EMERGENCY"]

    # Create the comparison sheet
    dir_path = os.path.join(current_directory, result_directory)
    file_path = os.path.join(dir_path, final_file)
    workbook = load_workbook(file_path)
    sheet = workbook.create_sheet(title="Comparison")


    title = "Load Test Cases"
    sheet[f'A1'] = title
    sheet[f'A1'].font = Font(bold=True)

    sheet[f'A2'] = "Variation1: is between KShort and ECMP"
    sheet[f'A3'] = "Variation2: is between KShort and ECMP+SRv6"
    sheet[f'A4'] = "Variation3: is between ECMP and ECMP+SRv6"

    # Empty line
    sheet.append([""])
    
    # Create a block for each test case
    for i, test_case in enumerate(test_cases):
        # Get max line considering the previous test cases
        max_line = sheet.max_row + 1

        set_test_case_headers(sheet, test_case, max_line)
        set_comparasion_formulas(sheet, max_line)
        print("Seting values copy from other sheets")
        set_copied_values(sheet, test_case, max_line)

        # Insert 2 empty lines
        sheet.append([""])
        sheet.append([""])

    # Save the workbook
    workbook.save(file_path)

def configure_final_file():
    set_pkt_loss()
    set_fist_pkt_delay()
    set_caculations()
    set_INT_results()
    set_Emergency_calculation()
    set_Comparison_sheet()

def main():
    global args, client, results
    
    check_files_exist()

    # Read the CSV and SRv6 files
    for file_index, filename in enumerate(args.f):
        #print(f"file_index: {file_index}, Filename: {filename}")
        results = {}                                              #reset the results dictionary between files
        read_csv_files(filename)
        if args.SRv6_index is not None and file_index in args.SRv6_index:
            #Retrive the SRv6 logs data for the current file
            read_SRv6_log(file_index)

        export_results(filename)  
    
    configure_final_file()
    adjust_columns_width()
    
    client.close()


if __name__ == "__main__":
    parse_args()
    main()

