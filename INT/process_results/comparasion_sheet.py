import constants
from openpyxl import load_workbook

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font



def get_line_column_to_copy_from(sheet_to_copy_from_name, variable_number, dscp):
    line =None
    col = None

    workbook = load_workbook(constants.final_file_path)
    sheet_to_copy_from = workbook[sheet_to_copy_from_name]

    variable_name = constants.headers_lines[variable_number]

    pass_1_occurance = True          #there are 2 Lines on collumn A that have the same name
    if variable_number == 14:
        pass_1_occurance = False 

    # sheet_to_copy_from, get the line of the cell that contains the variable_name on collumn A and the collumn after it
    for row in sheet_to_copy_from.iter_rows(min_row=constants.last_line_raw_data[sheet_to_copy_from_name] + 1, max_row = sheet_to_copy_from.max_row, min_col=1, max_col=1):
        
        cell_e = sheet_to_copy_from[f"E{row[0].row}"]
        value = str(cell_e.value).strip()

        if value != str(dscp):  # Check if Wrong DSCP
            continue      

        if pass_1_occurance == False and row[0].value == "AVG 1º Packet Delay (nanoseconds)":
            pass_1_occurance = True
            continue

        if variable_number <= 9:
            if row[0].value == variable_name:
                # Get the next collumn letter of the cell that contains the variable_name
                line = row[0].row
                col = get_column_letter(row[0].column + 1)
                break
        elif variable_number == 10 or variable_number == 12:
            if row[0].value == "Mean":
                line = row[0].row
                if variable_number == 10:
                    col = get_column_letter(row[0].column + 1)
                else:
                    col = get_column_letter(row[0].column + 2)
                break
        elif variable_number == 11 or variable_number == 13:
            if row[0].value == "Standard Deviation":
                line = row[0].row
                if variable_number == 11:
                    col = get_column_letter(row[0].column + 1)
                else:
                    col = get_column_letter(row[0].column + 2)
                break
        elif variable_number == 14:
            if row[0].value == "AVG 1º Packet Delay (nanoseconds)":
                line = row[0].row
                col = get_column_letter(row[0].column + 3)
                break
        elif variable_number == 15:
            if row[0].value == "AVG Flow Delay (nanoseconds)":
                line = row[0].row
                col = get_column_letter(row[0].column + 3)
                break

    return line, col

def set_algorithm_headers(sheet, start_line):

    # Set the lines names
    sheet[f'A{start_line + 1}'] = constants.headers_lines[0]
    sheet[f'A{start_line + 2}'] = constants.headers_lines[1]
    sheet[f'A{start_line + 3}'] = constants.headers_lines[2]
    sheet[f'A{start_line + 4}'] = constants.headers_lines[3]
    sheet[f'A{start_line + 5}'] = constants.headers_lines[4]
    sheet[f'A{start_line + 6}'] = constants.headers_lines[5]
    sheet[f'A{start_line + 7}'] = constants.headers_lines[6]
    sheet[f'A{start_line + 8}'] = constants.headers_lines[7]
    sheet[f'A{start_line + 9}'] = constants.headers_lines[8]
    sheet[f'A{start_line + 10}'] = constants.headers_lines[9]
    sheet[f'A{start_line + 11}'] = constants.headers_lines[10]
    sheet[f'A{start_line + 12}'] = constants.headers_lines[11]
    sheet[f'A{start_line + 13}'] = constants.headers_lines[12]
    sheet[f'A{start_line + 14}'] = constants.headers_lines[13]


    # Set lines names in bold text
    sheet[f'A{start_line + 1}'].font = Font(bold=True)
    sheet[f'A{start_line + 2}'].font = Font(bold=True)
    sheet[f'A{start_line + 3}'].font = Font(bold=True)
    sheet[f'A{start_line + 4}'].font = Font(bold=True)
    sheet[f'A{start_line + 5}'].font = Font(bold=True)
    sheet[f'A{start_line + 6}'].font = Font(bold=True)
    sheet[f'A{start_line + 7}'].font = Font(bold=True)
    sheet[f'A{start_line + 8}'].font = Font(bold=True)
    sheet[f'A{start_line + 9}'].font = Font(bold=True)
    sheet[f'A{start_line + 10}'].font = Font(bold=True)
    sheet[f'A{start_line + 11}'].font = Font(bold=True)
    sheet[f'A{start_line + 12}'].font = Font(bold=True)
    sheet[f'A{start_line + 13}'].font = Font(bold=True)
    sheet[f'A{start_line + 14}'].font = Font(bold=True)

def set_comparasion_formulas(sheet, start_line):
    # Set the formulas to compare the results between the test cases
    for i in range(1, constants.num_values_to_compare_all_tests + 1 - 2):
        #print(sheet[f'A{start_line + i}'].value)
        sheet[f'E{start_line + i}'] = f'=IFERROR(ROUND((C{start_line + i} - B{start_line + i}) / ABS(B{start_line + i}) * 100, 2), 0)'
        sheet[f'F{start_line + i}'] = f'=IFERROR(ROUND((D{start_line + i} - B{start_line + i}) / ABS(B{start_line + i}) * 100, 2), 0)'
        sheet[f'G{start_line + i}'] = f'=IFERROR(ROUND((D{start_line + i} - C{start_line + i}) / ABS(C{start_line + i}) * 100, 2), 0)'

def set_copied_values(sheet, current_test_scenario, start_line, dscp):    
    print("Seting values to copy from other sheets")
    
    # Cycle through the variables to compare (lines)
    for variable_number in range(constants.num_values_to_compare_all_tests - 2):
        
        # Cycle through the algorithms to compare (columns)
        for i in range(len(constants.algorithms)):
            curent_algorithm = constants.algorithms[i]

            # Get the name of the sheet to copy from
            sheet_to_copy_from_name = current_test_scenario + "-" + curent_algorithm
            print(f"For {current_test_scenario} DSCP:{dscp} copying variable nº {variable_number} from sheet: {sheet_to_copy_from_name}")

            line, column = get_line_column_to_copy_from(sheet_to_copy_from_name, variable_number, dscp)

            if line is None or column is None:
                print(f"Error getting line and column to copy from, sheet_to_copy_from: {sheet_to_copy_from_name}, variable number: {variable_number}")
                continue

            cell_reference = f"{column}{line}"
            formula = f"='{sheet_to_copy_from_name}'!{cell_reference}"
            sheet[f'{get_column_letter(2 + i)}{start_line + variable_number + 1}'] = formula

def set_scenario_headers(sheet, test_case, start_line):
    # Set test case name in bold test
        title = f"{test_case}"
        sheet[f'A{start_line}'] = title
        sheet[f'A{start_line}'].font = Font(bold=True)

        # Set the collumn names
        sheet[f'B{start_line}'] = constants.algorithms[0]
        sheet[f'C{start_line}'] = constants.algorithms[1]
        sheet[f'D{start_line}'] = constants.algorithms[2]
        sheet[f'E{start_line}'] = "Variation 1 (%)"
        sheet[f'F{start_line}'] = "Variation 2 (%)"
        sheet[f'G{start_line}'] = "Variation 3 (%)"

        # Set collumn names in bold text
        sheet[f'B{start_line}'].font = Font(bold=True)
        sheet[f'C{start_line}'].font = Font(bold=True)
        sheet[f'D{start_line}'].font = Font(bold=True)
        sheet[f'E{start_line}'].font = Font(bold=True)
        sheet[f'F{start_line}'].font = Font(bold=True)
        sheet[f'G{start_line}'].font = Font(bold=True)

def comparasion_area(sheet, current_test_scenario, start_line, dscp):
    #as bold text
    if dscp == -1:
        sheet[f'A{start_line}'] = "All DSCP: All Data Flows"
    else:
        sheet[f'A{start_line}'] = f"DSCP: {dscp}"
    sheet[f'A{start_line}'].font = Font(bold=True)

    start_line= start_line + 1

    set_algorithm_headers(sheet, start_line)
    set_comparasion_formulas(sheet, start_line)
    set_copied_values(sheet, current_test_scenario, start_line, dscp)
    sheet.append([""])

def set_Non_to_Emergency_Data_Flows_Comparasion(sheet, current_test_scenario, start_line):
    sheet[f'A{start_line}'] = "For All Data Flows"
    sheet[f'A{start_line + 1}'] = constants.headers_lines[-2]
    sheet[f'A{start_line + 2}'] = constants.headers_lines[-1]

    sheet[f'A{start_line}'].font     = Font(bold=True)
    sheet[f'A{start_line + 1}'].font = Font(bold=True)
    sheet[f'A{start_line + 2}'].font = Font(bold=True)

    for i in range(len(constants.algorithms)):
        sheet_to_copy_from_name = f"{current_test_scenario}-{constants.algorithms[i]}"

        line1, column1 = get_line_column_to_copy_from(sheet_to_copy_from_name, 14, -1)
        line2, column2 = get_line_column_to_copy_from(sheet_to_copy_from_name, 15, -1)

        if line1 is None or column1 is None:
            print(f"Error getting line and column to copy from, sheet_to_copy_from: {sheet_to_copy_from_name}, variable number: {14}")
            continue
        if line2 is None or column2 is None:
            print(f"Error getting line and column to copy from, sheet_to_copy_from: {sheet_to_copy_from_name}, variable number: {15}")
            continue

        cell_reference1 = f"{column1}{line1}"
        formula1 = f"='{sheet_to_copy_from_name}'!{cell_reference1}"
        
        cell_reference2 = f"{column2}{line2}"
        formula2 = f"='{sheet_to_copy_from_name}'!{cell_reference2}"

        sheet[f'{get_column_letter(2 + i)}{start_line + 1}'] = formula1
        sheet[f'{get_column_letter(2 + i)}{start_line + 2}'] = formula2

    #Comparasions SHOULD BE TAKEN OUT THIS FUNCTION AND THE LOOP THAT IS CONTAINING IT
    sheet[f'E{start_line + 1}'] = f'=IFERROR(ROUND((C{start_line + 1} - B{start_line + 1}) / ABS(B{start_line + 1}) * 100, 2), 0)'
    sheet[f'E{start_line + 2}'] = f'=IFERROR(ROUND((C{start_line + 2} - B{start_line + 2}) / ABS(B{start_line + 2}) * 100, 2), 0)'

    sheet[f'F{start_line + 1}'] = f'=IFERROR(ROUND((D{start_line + 1} - B{start_line + 1}) / ABS(B{start_line + 1}) * 100, 2), 0)'
    sheet[f'F{start_line + 2}'] = f'=IFERROR(ROUND((D{start_line + 2} - B{start_line + 2}) / ABS(B{start_line + 2}) * 100, 2), 0)'

    sheet[f'G{start_line + 1}'] = f'=IFERROR(ROUND((D{start_line + 1} - C{start_line + 1}) / ABS(C{start_line + 1}) * 100, 2), 0)'
    sheet[f'G{start_line + 2}'] = f'=IFERROR(ROUND((D{start_line + 2} - C{start_line + 2}) / ABS(C{start_line + 2}) * 100, 2), 0)'

def set_SRv6_area(sheet, current_test_scenario):
    sheet_target = current_test_scenario + "-ECMP-SRv6"
    last_line = sheet.max_row

    sheet[f'A{last_line + 1}'] = "SRv6 Rules"
    sheet[f'A{last_line + 2}'] = "AVG Nº of SRv6 rules Created"
    sheet[f'A{last_line + 3}'] = "AVG Nº of SRv6 rules Removed"
    sheet[f'B{last_line + 1}'] = "Values"

    sheet[f'A{last_line + 1}'].font = Font(bold=True) 
    sheet[f'A{last_line + 2}'].font = Font(bold=True)
    sheet[f'A{last_line + 3}'].font = Font(bold=True)
    sheet[f'B{last_line + 1}'].font = Font(bold=True)

    sheet[f'B{last_line + 2}'] = f'=COUNTIF(\'{sheet_target}\'!B1:B{constants.last_line_raw_data[sheet_target]}, \"Created SRv6 rule") / {constants.args.num_iterations}'
    sheet[f'B{last_line + 3}'] = f'=COUNTIF(\'{sheet_target}\'!B1:B{constants.last_line_raw_data[sheet_target]}, \"Removed SRv6 rule") / {constants.args.num_iterations}'



def set_Comparison_sheet():
    print("Setting the Comparison sheet")

    # Create the comparison sheet
    workbook = load_workbook(constants.final_file_path)
    sheet = workbook.create_sheet(title="Comparison")

    title = "Load Test Cases"
    sheet[f'A1'] = title
    sheet[f'A1'].font = Font(bold=True)

    sheet[f'A2'] = "Variation1: is betwee KShort and ECMP"
    sheet[f'A3'] = "Variation2: is betwee KShort and ECMP+SRv6"
    sheet[f'A4'] = "Variation3: is betwee ECMP and ECMP+SRv6"

    sheet[f'A2'].font = Font(bold=True)
    sheet[f'A3'].font = Font(bold=True)
    sheet[f'A4'].font = Font(bold=True)

    # Empty line
    sheet.append([""])
    
    # Create a block for each test case
    for current_test_scenario in constants.test_scenarios:
        print(f"Setting the Comparison sheet for test scenario: {current_test_scenario}")
        # Get max line considering the previous test cases
        max_line = sheet.max_row + 2
        set_scenario_headers(sheet, current_test_scenario, max_line)
        max_line = sheet.max_row + 1
        has_emergency_dscp = False

        for dscp in constants.DSCP_per_scenario[current_test_scenario]:
            if not has_emergency_dscp and dscp >= 40:
                has_emergency_dscp = True
            
            max_line = sheet.max_row + 1
            comparasion_area(sheet, current_test_scenario, max_line, dscp)

        # Set SRv6 area
        if constants.args.SRv6_index is not None:
            set_SRv6_area(sheet, current_test_scenario)

        # If contains DSCP >= 40, set comparasion for Non to Emergency Data Flows
        if has_emergency_dscp:
            sheet.append([""])
            set_Non_to_Emergency_Data_Flows_Comparasion(sheet, current_test_scenario, sheet.max_row + 1)

        # Insert 2 empty lines
        sheet.append([""])
        sheet.append([""])

    # Save the workbook
    workbook.save(constants.final_file_path)
