import argparse
import ast
import csv
import json
import os
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


import constants, export, configure

def adjust_columns_width():
    print(f"Adjusting columns width for all sheets")

    #open the workbook
    workbook = load_workbook(constants.final_file_path)

    # Adjust column widths to fit the text
    for sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]
        for column_cells in sheet.columns:
            length = max(len(str(cell.value).strip()) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length
    
    # Save the workbook
    workbook.save(constants.final_file_path)

def read_json(file_path):
    """
    Reads a JSON file and returns its content as a dictionary.
    
    :param file_path: Path to the JSON file.
    :return: Dictionary containing the JSON data.
    """
    with open(file_path, 'r') as file:
        data = json.load(file)
    return data

def get_pkt_size_dscp(flow):
    #reads the INT DB and sets the pkt size and DSCP collumns

    query = f"""
        SELECT dscp, size
        FROM flow_stats
        WHERE   "src_ip" = '{flow[0]}'
        AND     "dst_ip" = '{flow[1]}'
        AND     "flow_label" = '{flow[2]}'
        ORDER BY time DESC
        LIMIT 1
    """
    #print(f"Query: {query}")
    
    r = constants.apply_query(query)

    if r.raw["series"] == []:
        print(f"At get_pkt_size_dscp() Flow {flow} not found in the DB, probably multicast related")
        return -1, -1

    dscp = r.raw["series"][0]["values"][0][1]
    size = r.raw["series"][0]["values"][0][2]

    #print(dscp)
    #print(size)

    return dscp, size

def read_raw_results(row):
    iteration = row[0]
    ħost = row[1]
    flow = (row[2], row[3], int(row[4]))
    Is = row[5]
    number_of_packets = int(row[6])
    first_packet_time = row[7]
    dscp = int(row[10])
    values_end_points = {}
    values_end_points["num_pkt"] = number_of_packets
    values_end_points["time"] = float(first_packet_time)
    values_end_points["num_hosts"] = 1

    if Is == "receiver":
        number_out_of_order_packets = int(row[8])
        out_of_order_packets = row[9]
        avg_jitter = float(row[11])

        extra1 = {"num_out_of_order_pkt": number_out_of_order_packets}
        extra2 = {"out_of_order_pkt": out_of_order_packets}
        extra3 = {"avg_jitter": avg_jitter}

        values_end_points["extra"] = {}
        values_end_points["extra"].update(extra1)
        values_end_points["extra"].update(extra2)
        values_end_points["extra"].update(extra3)

    not_needed_anymore, pkt_size = get_pkt_size_dscp(flow)            #Get the flows info
    values_flow = {Is: values_end_points, "DSCP": dscp, "Packet Size": pkt_size}

    # Check if the iteration is already in the results dictionary
    if iteration not in constants.results:
        values_iteration = {flow: values_flow}
        constants.results[iteration] = values_iteration
    else:
        # Check if the flow is already in the constants.results dictionary
        if flow not in constants.results[iteration]:
            constants.results[iteration][flow] = values_flow
        else:
            # Add currect Is to the flow
            #Just add if sender or receiver no receiver there yet
            if Is == "sender" or (Is == "receiver" and "receiver" not in constants.results[iteration][flow]):      
                constants.results[iteration][flow][Is] = values_end_points

            else:                           #If multiple receivers, acommudate for that with averages amd concatenations
                old_number_hosts = constants.results[iteration][flow][Is]["num_hosts"]
                constants.results[iteration][flow][Is]["time"]    = (constants.results[iteration][flow][Is]["time"]    * old_number_hosts + values_end_points["time"]   ) / (old_number_hosts + 1)
                constants.results[iteration][flow][Is]["num_pkt"] = (constants.results[iteration][flow][Is]["num_pkt"] * old_number_hosts + values_end_points["num_pkt"]) / (old_number_hosts + 1)

                constants.results[iteration][flow][Is]["extra"]["num_out_of_order_pkt"] = (constants.results[iteration][flow][Is]["extra"]["num_out_of_order_pkt"] * old_number_hosts + number_out_of_order_packets) / (old_number_hosts + 1)
                constants.results[iteration][flow][Is]["extra"]["out_of_order_pkt"]    += out_of_order_packets
                constants.results[iteration][flow][Is]["extra"]["avg_jitter"]           = (constants.results[iteration][flow][Is]["extra"]["avg_jitter"] * old_number_hosts + avg_jitter) / (old_number_hosts + 1)

                constants.results[iteration][flow][Is]["num_hosts"] = old_number_hosts + 1

def read_csv_files(filename):
    first_row = True

    file_path = os.path.join(constants.results_path, filename)
    
    print(f"Reading files: {filename}")
    with open(file_path, 'r', newline='') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            #Skip the header, first row
            if first_row:
                first_row = False
                continue
            read_raw_results(row)

    #print("Done reading file")

def check_files_exist():
    # Check if the directory/files exist

    if not os.path.isdir(constants.results_path):  # Correct condition to check if the directory does not exist
        print(f"Directory {constants.result_directory} does not exist.")
        sys.exit(1)

    for filename in constants.args.f:
        file_path = os.path.join(constants.results_path, filename)
        # Check if the file exists
        if not os.path.isfile(file_path):
            print(f"File {filename} not found in {file_path}")
            sys.exit(1)

    #---------------Check if the SRv6 logs exist

    full_analy_path = os.path.join(constants.parent_path, constants.analyzer_directory) 

    # Check if the directory exists
    if not os.path.isdir(full_analy_path):
        print(f"Directory {constants.analyzer_directory} does not exist.")
        sys.exit(1)
    
    # Check if the log files exist
    if constants.args.SRv6_logs is not None:
        for log_file in constants.args.SRv6_logs:
            log_file_path = os.path.join(full_analy_path, log_file)
            #print(f"Checking SRv6 log file: {log_file_path}")

            if not os.path.isfile(log_file_path):
                print(f"File {log_file} not found in {log_file_path}")
                sys.exit(1)

def read_SRv6_line(line):
    #print(f"Reading SRv6 line: {line}")

    new_content = {}
    part1 = line.split(" - ")
    part2 = part1[2].split(" => ")
    operation = part2[0]
    part3 = part2[1].split(": ")
    iteration = part1[0]
    timestamp = part1[1]
    operation = part2[0]
    responsible_switch = int(part3[0])
    rule_elemets_str = line.split("{")[1]
    rule_elemets = ast.literal_eval("{"+rule_elemets_str)

    # Convert numebrs in the dictionary to integers
    rule_elemets["deviceID"] = int(rule_elemets["deviceID"])
    rule_elemets["flow_label"] = int(rule_elemets["flow_label"])
    rule_elemets["src_mask"] = int(rule_elemets["src_mask"])
    rule_elemets["dst_mask"] = int(rule_elemets["dst_mask"])
    rule_elemets["flow_label_mask"] = int(rule_elemets["flow_label_mask"])

    # Extract some values from the dictionary
    src_mask = rule_elemets["src_mask"]
    dst_mask = rule_elemets["dst_mask"]
    flow_label_mask = rule_elemets["flow_label_mask"]

    if src_mask != 128 or dst_mask != 128 or flow_label_mask != 255:
        print(f"Error in the masks in the rule: {rule_elemets}, we only work with flow specific rules")
        sys.exit(1)


    #---------------Create the new operation for the dictionary
    new_content["responsible_switch"] = responsible_switch    # Which switch the load induced the operation
    new_content["operation"] = operation
    new_content["timestamp"] = timestamp
    new_content["rule"] = rule_elemets

    # Add the results to the dictionary
    # At this point the iteration is already in the dictionary
    results_ite = constants.results[iteration]
    
    # Check if the operation if "SRv6_Operations" is already in the current iteration
    if "SRv6_Operations" not in results_ite:
        results_ite["SRv6_Operations"] = [new_content]   # as a list since there can be multiple operations
    else:
        # Add the operation to the dictionary
        results_ite["SRv6_Operations"].append(new_content)


def read_SRv6_log(file_index):
    analyzer_logs_dir = os.path.join(constants.parent_path, constants.analyzer_directory)
    
    #Get the position of the given file_index in args.SRv6_index
    log_index = constants.args.SRv6_index.index(file_index)

    # Read the SRv6 logs
    log_file = constants.args.SRv6_logs[log_index]
    log_file_path = os.path.join(analyzer_logs_dir, log_file)

    print(f"Reading SRv6 logs from file: {log_file}")

    # Read the SRv6 logs
    with open(log_file_path, 'r') as file:
        for line in file:
            read_SRv6_line(line)

def parse_args():

    parser = argparse.ArgumentParser(description='process parser')
    parser.add_argument('--f', help='CSV files to be processed',
                        type=str, action="store", required=True, nargs='+')
    
    # 2 arguments with multiple values, the nº of elements must be the same between them and the files
    parser.add_argument('--start', help='Timestamp (RFC3339 format) of when each test started (1 peer file)',
                        type=str, action="store", required=True, nargs='+')
    parser.add_argument('--end', help='Timestamp (RFC3339 format) of when each test ended (1 peer file)',
                        type=str, action="store", required=True, nargs='+')
    
    # 3 Optional argument, that must be used simultaneously
    parser.add_argument('--SRv6_index', help='Indexs of the files that have SRv6 logs associated to them (starting from 0)',
                        type=int, action="store", required=False, nargs='+')
    parser.add_argument('--SRv6_logs', help='Names of the log files with the SRv6 rules from previous argument (must match the order)',
                        type=str, action="store", required=False, nargs='+')
    parser.add_argument('--num_iterations', help='Nº of iterations on every test, help SRv6 AVG calculations)',
                        type=int, action="store", required=False)

    constants.args = parser.parse_args()
    
    # Check if the number of elements start/end is the same
    if len(constants.args.start) != len(constants.args.end):
        parser.error("The number of elements in --start and --end must be the same")
    if len(constants.args.start) != len(constants.args.f):
        parser.error("The number of elements in --start and --end must be the same as the number of files")

    # Check if constants.args pairs SRv6_index, SRv6_logs and num_iterations are used together
    if (constants.args.SRv6_index and not constants.args.SRv6_logs) or (not constants.args.SRv6_index and constants.args.SRv6_logs):
        parser.error("Both --SRv6_index and --SRv6_logs must be used together")
    if not constants.args.num_iterations and constants.args.SRv6_index:
        parser.error("The --num_iterations must be used with --SRv6_index and --SRv6_logs")

    # Check if the number of elements is the same in SRv6_index and SRv6_logs
    if constants.args.SRv6_index and constants.args.SRv6_logs:
        if len(constants.args.SRv6_index) != len(constants.args.SRv6_logs):
            parser.error("The number of elements in --SRv6_index and --SRv6_logs must be the same")

    # If index of SRv6 is given, check if it is valid
    if constants.args.SRv6_index:
        for index in constants.args.SRv6_index:
            if index < 0 or index >= len(constants.args.f):
                parser.error("The SRv6_index: "+ str(index) +" is invalid. It must be between 0 and the number of files-1")

def main():
    # In constants.args.f get for each element between - and 1ª _
    # No duplicated algorithms values
    seen = set()
    constants.algorithms = []
    for x in constants.args.f:
        key = x.split("-", 1)[1]
        key = key.split("_", 1)[0]
        if key not in seen:
            seen.add(key)
            constants.algorithms.append(key)

    # No duplicated test_scenarios values
    seen = set()
    constants.test_scenarios = []
    for x in constants.args.f:
        key = x.split("-")[0]
        if key not in seen:
            seen.add(key)
            constants.test_scenarios.append(key)

    #constants.DSCP_IPs = read_json(constants.filename_with_sizes)
    #print("Packet DSCP Multicast IPs read:\n",constants.DSCP_IPs)

    check_files_exist()

    # Delete the final file if it exists
    if os.path.isfile(constants.final_file_path):
        os.remove(constants.final_file_path)

    # Read the CSV files
    for file_index, filename in enumerate(constants.args.f):
        constants.results = {}                                              #reset the results dictionary between files

        read_csv_files(filename)
        if constants.args.SRv6_index is not None and file_index in constants.args.SRv6_index:
            #Retrive the SRv6 logs data for the current file
            read_SRv6_log(file_index)

        constants.calulate_std_jitter_per_dscp(filename)                    #calculate the std from jitter per DSCP
        export.export_raw_results(filename)                                 #export the results to the final file
    
    configure.configure_final_file()
    adjust_columns_width()
    
    constants.client.close()

if __name__ == "__main__":
    parse_args()
    main()

