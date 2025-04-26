import os
import pprint
import sys
import tempfile
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from PIL import Image as PILImage, ImageDraw, ImageFont
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter


import constants

def create_CDF(image_path, data, labels, xlabel, ylabel, x_min, x_max, title):
    # Create directory if it doesn't exist
    if not os.path.exists(constants.directory_images):
        os.makedirs(constants.directory_images)

    plt.figure(figsize=(6, 4))

    # Loop through each dataset and plot its CDF
    for i, dataset in enumerate(data):
        sorted_data = np.sort(dataset)
        cdf = np.arange(1, len(dataset) + 1) / len(dataset)
        plt.plot(sorted_data, cdf, marker='.', linestyle='solid', label=labels[i])

    # Graph formatting
    plt.title(title, fontsize=20)
    plt.xlabel(xlabel, fontsize=20)
    plt.ylabel(ylabel, fontsize=20)
    plt.xlim(x_min, x_max)
    plt.tick_params(axis='both', which='major', labelsize=14)
    plt.grid(True)
    plt.legend(fontsize=14)
    plt.tight_layout()
    plt.savefig(image_path)  # Save the image
    plt.close()  # Close the figure to free up memory


def create_CDF_graphs(sheet, datas, title, xlabel, ylabel, variable_name, position_image_x, position_image_y):
    print(f"Creating CDF with title: {title}...")
    image_path = constants.images_path + "/" + title + ".png"
    #remove from the path any presnece of ":"
    image_path = image_path.replace(":", "")

    # Concatnate all data to get global min and max
    all_data_contatenated = np.concatenate(datas).astype(np.float64)
    if all_data_contatenated.size == 0:
        print(f"Warning: No data available for {title}. Skipping CDF creation.")
        return None
    
    x_min, x_max = all_data_contatenated.min(), all_data_contatenated.max()
    create_CDF(image_path, datas, constants.algorithms, xlabel, ylabel, x_min, x_max, title)

    '''
    image_paths = []

    for index_current_algorithm, current_algorithm_name in enumerate(constants.algorithms):
        image_path = constants.images_path + "/" + current_algorithm_name + "-" + variable_name + ".png"
        current_data = np.array(datas[index_current_algorithm], dtype=np.float64)
        create_CDF(image_path, current_data, xlabel, ylabel, x_min, x_max, current_algorithm_name)
        image_paths.append(image_path)


    #-----------------------------------------------------------Create a combined image with title
    # Open all images and calculate total width and max height
    images = [PILImage.open(img_path) for img_path in image_paths]
    total_width = sum(img.width for img in images)
    max_height = max(img.height for img in images)
    title_height = 40  # space for title at the top

    # Create a new blank image using PILImage (from PIL), ADDING title_height to height
    image_collection = PILImage.new('RGB', (total_width, max_height + title_height), color='white')

    # Draw title
    draw = ImageDraw.Draw(image_collection)
    font = ImageFont.truetype("DejaVuSans.ttf", 35)  # Supports º and other extended Unicode characters

    bbox = draw.textbbox((0, 0), title, font=font)
    text_width  = bbox[2] - bbox[0]
    text_height = bbox[3] - bbox[1]
    draw.text( ((total_width - text_width) // 2, (title_height - text_height) // 2), title, fill="black", font=font)

    # Paste each image BELOW the title area
    current_x = 0
    for img in images:
        image_collection.paste(img, (current_x, title_height))
        current_x += img.width

    # Save the combined image to a temporary file
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_file:
        temp_file_path = temp_file.name
        image_collection.save(temp_file_path)
    '''

    # Create an openpyxl Image object from the temporary file
    openpyxl_image = Image(image_path)
    openpyxl_image.anchor = f'{get_column_letter(position_image_x)}{position_image_y}'  # Position the image in the sheet
    sheet.add_image(openpyxl_image)

    return 0

def from_excel_data(sheet, current_scenario, position_image_x, position_image_y):
    x_label = ""
    for current_collunm_index in constants.index_of_headers_to_do_CDF_out_of_raw_values:            
        datas = []
        x_label = constants.units_for_each_index_collumn[current_collunm_index]
        variable_title = constants.title_for_each_index_collumn[current_collunm_index]

        # Iterate through algorithms
        for current_algorithm in constants.algorithms:
            current_sheet_name = current_scenario + "-" + current_algorithm

            # Read the Excel file. VERY CONSUMING
            current_data = pd.read_excel(constants.final_file_path, sheet_name=current_sheet_name, header=None)

            # Get collumn by index
            current_data = current_data.iloc[:, current_collunm_index]
            current_data = current_data.dropna().values                       # Remove empty values
            variable_name = current_data[0]                                   # Get the first value of the column

            current_data = current_data[1:]                                   # Remove the first row (header)

            # Apply percentile
            percentile_value = np.percentile(current_data, constants.percentile)
            current_data = current_data[current_data <= percentile_value]      # Filter the values with the percentile value
            
            datas.append(current_data)

        # Create CDF graphs
        title = f"{current_scenario}: {variable_title}"
        res = create_CDF_graphs(sheet, datas, title, x_label, "Probability", variable_name, position_image_x, position_image_y)
        if res is not None:
            position_image_x += 10      # Move right for each algorithm
    
    return position_image_x

def from_db_data(sheet, current_scenario, position_image_x, position_image_y):
    x_label = "Nanoseconds"

    for current_table, current_values in constants.variables_to_do_CDF_out_of_db_values.items():    
        for current_variable in current_values: 
            if current_table == "flow_stats":
                title1 = "Flow" + " " + current_variable
            elif current_table == "switch_stats":
                title1 = "Hop" + " " + current_variable

            datas = []

            # Iterate through algorithms
            for current_algorithm in constants.algorithms:
                current_sheet_name = current_scenario + "-" + current_algorithm

                start_time, end_time = constants.start_end_times[current_sheet_name]

                # Read the DB data
                current_data = constants.get_full_variable_data_from_db(current_variable, constants.percentile, current_table, start_time, end_time)
                datas.append(current_data)

            # Create CDF graphs
            title = f"{current_scenario}: {title1}"
            res = create_CDF_graphs(sheet, datas, title, x_label, "Probability", current_variable, position_image_x, position_image_y)
            if res is not None:
                position_image_x += 10      # Move right for each algorithm
    
    return position_image_x

def create_graphs():
    workbook = load_workbook(constants.final_file_path)
    sheet = workbook.create_sheet(title="CDF Plots")
    position_image_y = 1

    # Get data and create graph for 3 algorithms for each scenario for each variable
    for current_scenario in constants.test_scenarios:
        print(f"Creating CDF plots for scenario {current_scenario}...")
        position_image_x = 1
        position_image_x = from_excel_data(sheet, current_scenario, position_image_x, position_image_y)
        position_image_x = from_db_data(sheet, current_scenario, position_image_x, position_image_y)
        position_image_y += 23          # Move down for each scenario 

    # Save the workbook
    workbook.save(constants.final_file_path)

