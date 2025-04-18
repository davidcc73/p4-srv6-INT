import os
import pprint
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

import constants



def extract_Is_values(line, iteration, flow, Is):
    line = line + [Is]
    # Is values, all in the current line
    for key, Is_value in constants.results[iteration][flow][Is].items():
        if key == "extra":
            for key, value in Is_value.items():
                line = line + [value]
        else:
            if key == "num_hosts":          #skip the num_hosts for final file
                continue
            line = line + [Is_value]

    return line

def export_SRv6_rules(sheet, iteration):
    sheet.append([""])

    # Add the "SRv6 Operations" header
    header = ["SRv6 Operations"]
    new_next_row = sheet.max_row + 1
    for col_num, value in enumerate(header, 1):
        cell = sheet.cell(row=new_next_row, column=col_num, value=value)
        cell.font = Font(bold=True)
    
    # Add the detailed headers for SRv6 operations
    header = ["Timestamp", "Operation", "Responsible Switch", "Source", "Destination", "Flow Label"]
    new_next_row = sheet.max_row + 1
    for col_num, value in enumerate(header, 1):
        cell = sheet.cell(row=new_next_row, column=col_num, value=value)
        cell.font = Font(bold=True)

    # Add the SRv6 operations
    for operation in constants.results[iteration]["SRv6_Operations"]:
        new_next_row = sheet.max_row + 1
        rule_src_IP = operation["rule"]["srcIP"]
        rule_dst_IP = operation["rule"]["dstIP"]
        rule_flow_label = operation["rule"]["flow_label"]
        line = [operation["timestamp"], operation["operation"], operation["responsible_switch"], rule_src_IP, rule_dst_IP, rule_flow_label]
        sheet.append(line)

    sheet.append([""])

def export_raw_results(OG_file):
    # Get the sheet name from filename before (_)
    sheet_name = OG_file.split("_")[0]
    
    # Check if the file exists
    file_path = os.path.join(constants.results_path, constants.final_file)
    if os.path.exists(file_path):
        # Load the existing workbook
        workbook = load_workbook(file_path)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(title=sheet_name)
    else:
        # Create a new workbook and sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
    
    # Write the header
    header = ["Flow src", "Flow dst", "Flow Label", "DSCP", "Packet Size (Bytes)", "Is", "Nº of packets", "1º Packet Timestamp(seconds)", "Nº of out of order packets", "Out of order packets", "AVG Flow Jitter (nanoseconds)"]
    for col_num, value in enumerate(header, 1):
        cell = sheet.cell(row=1, column=col_num, value=value)
        cell.font = Font(bold=True)

    # Write results, iteration by iteration
    for iteration in constants.results:
        
        # Write key
        sheet.append([f""])
        sheet.append([None])  # This adds an empty row
        cell = sheet.cell(row=sheet.max_row, column=1)  # Get the last row and column 1
        cell.value = f"Iteration - {iteration}"
        cell.font = Font(bold=True)

        # Flow by flow
        for flow in constants.results[iteration]:
            if flow == "SRv6_Operations":                                #skip the SRv6_Operations data for the current iteration
                continue
            
            keys = list(constants.results[iteration][flow].keys())

            DSCP = constants.results[iteration][flow]["DSCP"]
            pkt_size =  constants.results[iteration][flow]["Packet Size"]

            # Check if both types of Is are keys in the dictionary
            if "sender" in keys:
                OG_line = list(flow) + [DSCP, pkt_size]
                line_s = extract_Is_values(OG_line, iteration, flow, "sender")    
            else:
                print(f"Sender not found in iteration {iteration} flow {flow}")
                # empty tuple
                line_s = ()


            if "receiver" in keys:
                OG_line = list(flow) + [DSCP, pkt_size]
                line_r = extract_Is_values(OG_line, iteration, flow, "receiver")
            else:
                print(f"Receiver not found in iteration {iteration} flow {flow}")
                line_r = ()

            # Is by Is, sender must be the 1º
            sheet.append(line_s)
            sheet.append(line_r)
        
        # Write the SRv6 operations of this Iteration if they exist
        if "SRv6_Operations" in constants.results[iteration]:
            export_SRv6_rules(sheet, iteration)

        # Store the last line of raw data for the current sheet
        constants.last_line_raw_data[sheet_name] = sheet.max_row

    # Save the workbook
    workbook.save(file_path)
