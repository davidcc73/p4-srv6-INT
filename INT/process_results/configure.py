from cmath import sqrt
import os
import sys
import constants, comparasion_sheet, graphs
from openpyxl import load_workbook
from openpyxl.styles import Font

def get_byte_sum(start, end, dscp, dscp_condition):
    # Initialize the result dictionary to store total byte counts per switch ID
    sum = {}
    switch_ids = list(range(1, constants.num_switches + 1))

    # Loop through each unique switch ID
    for switch_id in switch_ids:
        # Formulate the query to get the sum of bytes for the given switch ID and time range
        percentile_query = f"""
            SELECT PERCENTILE("size", {constants.percentile}) AS p_size
            FROM "flow_stats"
            WHERE time >= '{start}'
            AND time <= '{end}'
        """
        percentile_value = constants.apply_query(percentile_query)
        percentile_value = list(percentile_value.get_points())[0]['p_size']


        query = f"""                                    
            SELECT SUM("size") AS total_count
            FROM flow_stats 
            WHERE time >= '{start}' AND time <= '{end}' 
            AND path =~ /(^|-)({switch_id})(-|$|\b)/
            {dscp_condition}
            AND "size" <= {percentile_value}
        """

        result = constants.apply_query(query)  # Assume this returns a dictionary with 'total_count'
        
        # Add the result to the sum dictionary under the switch_id key and dscp key
        if dscp not in sum:
            sum[dscp] = {}
        if switch_id not in sum[dscp]:
            sum[dscp][switch_id] = {}
        
        if not result.raw["series"]:            #if there is no data, set to 0
            sum[dscp][switch_id]["Byte Sums"] = 0
        else:
            sum[dscp][switch_id]["Byte Sums"] = result.raw["series"][0]["values"][0][1]

    return sum

def calculate_percentages(start, end, switch_data, dscp, dscp_condition):
    percentile_query = f"""
        SELECT PERCENTILE("latency", {constants.percentile}) AS p_latency
        FROM "flow_stats"
        WHERE time >= '{start}'
        AND time <= '{end}'
    """

    flow_percentile_value = constants.apply_query(percentile_query)
    flow_percentile_value = list(flow_percentile_value.get_points())[0]['p_latency']

    # Get the total count of packets
    query = f"""
        SELECT COUNT("latency") AS total_count
        FROM flow_stats
        WHERE time >= '{start}' 
        AND time <= '{end}'
        {dscp_condition}
        AND "latency" <= {flow_percentile_value}
    """
    result = constants.apply_query(query)
    total_count = result.raw["series"][0]["values"][0][1]  # Extract total_count from the result

    ############ Get the count of packets that went to each switch
    percentile_query = f"""
        SELECT PERCENTILE("latency", {constants.percentile}) AS p_latency
        FROM "switch_stats"
        WHERE time >= '{start}'
        AND time <= '{end}'
    """

    switch_percentile_value = constants.apply_query(percentile_query)
    switch_percentile_value = list(switch_percentile_value.get_points())[0]['p_latency']

    query = f"""
        SELECT COUNT("latency") AS switch_count
        FROM switch_stats
        WHERE time >= '{start}'
        AND time <= '{end}'
        {dscp_condition}
        AND "latency" <= {switch_percentile_value}
        GROUP BY switch_id
    """
    result = constants.apply_query(query)
    
    # Calculate the percentage of packets that went to each switch
    # initialize to all switches as 0, so unused switches are taken into account too
    for switch_id in range(1, constants.num_switches + 1):
        switch_data[dscp][switch_id]["Percentage Pkt"] = 0

    for row in result.raw["series"]:
        #tuple pair: id, count
        switch_id = int(row["tags"]["switch_id"])
        switch_count = int(row["values"][0][1])
        switch_data[dscp][switch_id]["Percentage Pkt"] = round((switch_count / total_count) * 100, 2)

    return switch_data

def get_mean_standard_deviation(switch_data, dscp):
    sum_percentage = 0
    sum_byte = 0
    count = 0

    # Get the mean of the (Byte Sums) and (Percentage Pkt) for all switches in switch_data
    for switch_id in switch_data[dscp]:
        sum_percentage += switch_data[dscp][switch_id]["Percentage Pkt"]
        sum_byte       += switch_data[dscp][switch_id]["Byte Sums"]
        count += 1

    percentage_mean = sum_percentage / count
    byte_mean = sum_byte / count
    

    # Get the Standard Deviation of the (Byte Sums) and (Percentage Pkt) for all switches in switch_data
    sum_squared_diff_percentage = 0
    sum_squared_diff_byte = 0
    
    for switch_id in switch_data[dscp]:
        current_percentage = switch_data[dscp][switch_id]["Percentage Pkt"]
        current_byte       = switch_data[dscp][switch_id]["Byte Sums"]
        sum_squared_diff_percentage += (current_percentage - percentage_mean) ** 2
        sum_squared_diff_byte       += (current_byte - byte_mean) ** 2
    
    percentage_std_dev = sqrt(sum_squared_diff_percentage / count).real
    byte_std_dev = sqrt(sum_squared_diff_byte / count).real
    
    switch_data[dscp]["Percentage Mean"]               = round(percentage_mean, 2)
    switch_data[dscp]["Byte Mean"]                     = round(byte_mean, 2)
    switch_data[dscp]["Percentage Standard Deviation"] = round(percentage_std_dev, 2)
    switch_data[dscp]["Byte Standard Deviation"]       = round(byte_std_dev, 2)

    return switch_data

def write_INT_results_switchID(sheet, switch_data, dscp):
    # Write the results in the sheet
    last_line = sheet.max_row + 1

    # Set new Switch ID headers
    if dscp == -1:
        title = "Switch ID For All Flows"
    else:
        title = f"Switch ID For Flows with DSCP = {dscp}"

    sheet[f'A{last_line + 1}'] = title
    sheet[f'B{last_line + 1}'] = "% of packets to each switch"
    sheet[f'C{last_line + 1}'] = "Total Sum of Processed Bytes"
    sheet[f'E{last_line + 1}'] = "DSCP"

    sheet[f'A{last_line + 1}'].font = Font(bold=True)
    sheet[f'B{last_line + 1}'].font = Font(bold=True)
    sheet[f'C{last_line + 1}'].font = Font(bold=True)
    sheet[f'E{last_line + 1}'].font = Font(bold=True)


    # Write percentages and total bytes processed, cycle through keys that are numbers
    for i, switch_id in enumerate(switch_data[dscp].keys()):
        if isinstance(switch_id, int):                #skip sets that are non-switch_id
            sheet[f'A{last_line + 2 + i}'] = switch_id
            
            #percentage of total packets that went to each switch
            sheet[f'B{last_line + 2 + i}'] = switch_data[dscp][switch_id]["Percentage Pkt"]
            
            #Sum of processed bytes
            sheet[f'C{last_line + 2 + i}'] = switch_data[dscp][switch_id]["Byte Sums"]

            #Set auxiliar DSCP collumn
            sheet[f'E{last_line + 2 + i}'] = dscp


    # Write the mean and standard deviation of the percentages and bytes
    sheet[f'A{last_line + constants.num_switches + 1 + 1}'] = "Mean"
    sheet[f'A{last_line + constants.num_switches + 1 + 2}'] = "Standard Deviation"
    sheet[f'A{last_line + constants.num_switches + 1 + 1}'].font = Font(bold=True)
    sheet[f'A{last_line + constants.num_switches + 1 + 2}'].font = Font(bold=True)

    sheet[f'B{last_line + constants.num_switches + 1 + 1}'] = switch_data[dscp]["Percentage Mean"]
    sheet[f'B{last_line + constants.num_switches + 1 + 2}'] = switch_data[dscp]["Percentage Standard Deviation"]
    sheet[f'C{last_line + constants.num_switches + 1 + 1}'] = switch_data[dscp]["Byte Mean"]
    sheet[f'C{last_line + constants.num_switches + 1 + 2}'] = switch_data[dscp]["Byte Standard Deviation"]

    sheet[f'E{last_line + constants.num_switches + 1 + 1}'] = dscp
    sheet[f'E{last_line + constants.num_switches + 1 + 2}'] = dscp

def write_INT_results(sheet, AVG_flows_latency, STD_flows_latency, AVG_hop_latency, STD_hop_latency, dscp):
    # Write the results in the sheet
    last_line = sheet.max_row + 1

    # Set new Calculation headers
    sheet[f'A{last_line + 0}'] = "AVG Flows Latency (nanoseconds)"
    sheet[f'A{last_line + 1}'] = "STD Flows Latency (nanoseconds)"
    sheet[f'A{last_line + 2}'] = "AVG Hop Latency (nanoseconds)"
    sheet[f'A{last_line + 3}'] = "STD Hop Latency (nanoseconds)"

    sheet[f'A{last_line + 0}'].font = Font(bold=True)
    sheet[f'A{last_line + 1}'].font = Font(bold=True)
    sheet[f'A{last_line + 2}'].font = Font(bold=True)
    sheet[f'A{last_line + 3}'].font = Font(bold=True)

    sheet[f'B{last_line + 0}'] = AVG_flows_latency
    sheet[f'B{last_line + 1}'] = STD_flows_latency
    sheet[f'B{last_line + 2}'] = AVG_hop_latency
    sheet[f'B{last_line + 3}'] = STD_hop_latency

    sheet[f'E{last_line + 0}'] = dscp
    sheet[f'E{last_line + 1}'] = dscp
    sheet[f'E{last_line + 2}'] = dscp
    sheet[f'E{last_line + 3}'] = dscp

def set_pkt_loss():
    # To all sheets set pkt loss in the raw data area for all iterations

    # Configure each sheet  
    workbook = load_workbook(constants.final_file_path)

    # Set formula for each sheet
    for sheet in workbook.sheetnames:
        sheet = workbook[sheet]

        #Set new headers
        sheet['M1'] = "Packet Loss"
        sheet['N1'] = "Packet Loss (%)"

        sheet['M1'].font = Font(bold=True)
        sheet['N1'].font = Font(bold=True)

        no_formula_section = False

        for row in sheet.iter_rows(min_row=2, max_row=constants.last_line_raw_data[sheet.title], min_col=1, max_col=3):
            if row[0].value == "Calculations":
                break

            if row[0].value and row[0].value.startswith("Iteration -"):
                no_formula_section = False
            if no_formula_section:
                continue

            #if cell from collumn A does not contain an IPv6 address, skip
            if row[0].value is None or ":" not in row[0].value:
                skip = True
                continue
            if skip:            #not in the right line of the pair
                skip = False
                continue
            
            # get value from collumn G, line row[0].row-1
            pkts_received = sheet[f'G{row[0].row-1}'].value
            pkts_sent = sheet[f'G{row[0].row}'].value
            
            if pkts_received is None or pkts_sent is None:      #Skip the SRv6 area
                continue

            # Set the formula, pkt loss, -1 is sender, 0 is receiver
            pkt_loss = pkts_received - pkts_sent
            sheet[f'M{row[0].row}'] = pkt_loss   
            sheet[f'N{row[0].row}'] = round((pkt_loss/pkts_sent)*100, 2)

            skip = True

    # Save the workbook
    workbook.save(constants.final_file_path)

def set_fist_pkt_delay():
    # Configure each sheet
    workbook = load_workbook(constants.final_file_path)
    
    # Set formula for each sheet
    for sheet in workbook.sheetnames:
        sheet = workbook[sheet]

        #Set new headers as bold text
        sheet['O1'] = "1º Packet Delay (nanoseconds)"
        sheet['O1'].font = Font(bold=True)

        no_formula_section = False
        
        # Set collumn L to contain a formula to be the subtraction of values of collum F of the current pair of lines
        for row in sheet.iter_rows(min_row=2, max_row=constants.last_line_raw_data[sheet.title], min_col=1, max_col=3):
            if row[0].value == "Calculations":
                break

            if row[0].value and row[0].value.startswith("Iteration -"):
                no_formula_section = False
            if no_formula_section:
                continue

            #if cell from collumn A does not contain an IPv4 address, skip
            if row[0].value is None or ":" not in row[0].value:
                skip = True
                continue
            if skip:            #not in the right line of the pair
                skip = False
                continue
            #print(row)
            
            # Set the formula, pkt loss, -1 is sender, 0 is receiver
            # The values are 2 Timestamp (seconds-Unix Epoch)
            # subtraction give seconds, we convert to nanoseconds
            first_pkt_sent = sheet[f'H{row[0].row-1}'].value
            first_pkt_received = sheet[f'H{row[0].row}'].value

            if first_pkt_received is None or first_pkt_sent is None:      #Skip the SRv6 area
                continue

            first_pkt_delay = first_pkt_received - first_pkt_sent
            sheet[f'O{row[0].row}'] = round(first_pkt_delay*1000000000, 2)  

            skip = True

    # Save the workbook
    workbook.save(constants.final_file_path)

def set_caculation_formulas(workbook, sheet_name, dscp, scenario_DSCPs):

    if dscp == -1:
        title = "Calculations For All Flows"
    else:
        title = f"Calculations For Flows with DSCP = {dscp}"

    sheet = workbook[sheet_name]
    last_line_raw_data_sheet = constants.last_line_raw_data[sheet_name]

    #Pass the last line with data, and leave 2 empty lines
    last_line = sheet.max_row + 4

    #Set new headers
    sheet[f'A{last_line}'] = title
    sheet[f'A{last_line + 1}'] = "AVG Out of Order Packets (Nº)"
    sheet[f'A{last_line + 2}'] = "AVG Packet Loss (Nº)"
    sheet[f'A{last_line + 3}'] = "AVG Packet Loss (%)"
    sheet[f'A{last_line + 4}'] = "AVG 1º Packet Delay (nanoseconds)"
    sheet[f'A{last_line + 5}'] = "AVG Flow Jitter (nanoseconds)"
    sheet[f'A{last_line + 6}'] = "STD Flow Jitter (nanoseconds)"
    sheet[f'B{last_line}'] = "Values"
    sheet[f'E{last_line}'] = "DSCP"

    sheet[f'A{last_line}'].font = Font(bold=True)
    sheet[f'A{last_line + 1}'].font = Font(bold=True)
    sheet[f'A{last_line + 2}'].font = Font(bold=True)
    sheet[f'A{last_line + 3}'].font = Font(bold=True)
    sheet[f'A{last_line + 4}'].font = Font(bold=True)
    sheet[f'A{last_line + 5}'].font = Font(bold=True)
    sheet[f'A{last_line + 6}'].font = Font(bold=True)
    sheet[f'B{last_line}'].font = Font(bold=True)
    sheet[f'E{last_line}'].font = Font(bold=True)

    sheet[f'E{last_line + 1}'] = dscp
    sheet[f'E{last_line + 2}'] = dscp
    sheet[f'E{last_line + 3}'] = dscp
    sheet[f'E{last_line + 4}'] = dscp
    sheet[f'E{last_line + 5}'] = dscp
    sheet[f'E{last_line + 6}'] = dscp


    #-----------------------------------------------------------------------------------------------------Calculations
    avg_collunm_I = constants.get_collumn_average_per_dscp(sheet, last_line_raw_data_sheet, "D", dscp, "I", scenario_DSCPs)
    avg_collunm_M = constants.get_collumn_average_per_dscp(sheet, last_line_raw_data_sheet, "D", dscp, "M", scenario_DSCPs)
    avg_collunm_N = constants.get_collumn_average_per_dscp(sheet, last_line_raw_data_sheet, "D", dscp, "N", scenario_DSCPs)
    avg_collunm_O = constants.get_collumn_average_per_dscp(sheet, last_line_raw_data_sheet, "D", dscp, "O", scenario_DSCPs)
    avg_collunm_K = constants.get_collumn_average_per_dscp(sheet, last_line_raw_data_sheet, "D", dscp, "K", scenario_DSCPs)

    sheet[f'B{last_line + 1}'] = avg_collunm_I
    sheet[f'B{last_line + 2}'] = avg_collunm_M
    sheet[f'B{last_line + 3}'] = avg_collunm_N
    sheet[f'B{last_line + 4}'] = avg_collunm_O
    sheet[f'B{last_line + 5}'] = avg_collunm_K
    sheet[f'B{last_line + 6}'] = constants.aux_calculated_results[sheet_name][dscp]["std_jitter"]       #array formuals are not working, so we calculated and set the value here


def get_avg_stdev_flow_hop_latency(start, end, dscp_condition):
    ############################################ Get the results from the DB
    # We need AVG Latency of ALL flows combined (NOT distinguishing between flows)
    # Query to get the 95th percentile latency value, to exclude outliers
    percentile_query = f"""
        SELECT PERCENTILE("latency", {constants.percentile}) AS p_latency
        FROM flow_stats
        WHERE time >= '{start}'
        AND time <= '{end}'
        {dscp_condition}
    """
    percentile_result = constants.apply_query(percentile_query)
    p_latency = list(percentile_result.get_points())[0]['p_latency']                #nanoseconds

    query = f"""
                SELECT MEAN("latency"), STDDEV("latency")
                FROM  flow_stats
                WHERE time >= '{start}'
                AND time <= '{end}'
                AND "latency" <= {p_latency}
                {dscp_condition}
            """
    result = constants.apply_query(query)
    AVG_flows_latency = round(result.raw["series"][0]["values"][0][1], 2)           #nanoseconds
    STD_flows_latency = round(result.raw["series"][0]["values"][0][2], 2)

    ###########################################
    # We need AVG Latency for processing of ALL packets (NOT distinguishing between switches/flows) 
    # Query to get the 95th percentile latency value, to exclude outliers
    percentile_query = f"""
        SELECT PERCENTILE("latency", {constants.percentile}) AS p_latency
        FROM switch_stats
        WHERE time >= '{start}'
        AND time <= '{end}'
        {dscp_condition}
    """
    percentile_result = constants.apply_query(percentile_query)
    p_latency = list(percentile_result.get_points())[0]['p_latency']                #nanoseconds
    
    query = f"""
                SELECT MEAN("latency"), STDDEV("latency")
                FROM  switch_stats
                WHERE time >= '{start}'
                AND time <= '{end}'
                AND "latency" <= {p_latency}
                {dscp_condition}
            """
    result = constants.apply_query(query)
    AVG_hop_latency = round(result.raw["series"][0]["values"][0][1], 2)             #nanoseconds
    STD_hop_latency = round(result.raw["series"][0]["values"][0][2], 2)         

    return AVG_flows_latency, STD_flows_latency, AVG_hop_latency, STD_hop_latency

def set_INT_results(workbook, sheet_name, dscp, i):

    #can i can not exceed the number of args.f (last one is comparasions)
    if i >= len(constants.args.f):
        return

    if dscp == -1:
        dscp_condition = ""
    else:
        dscp_condition = f"AND dscp = \'{dscp}\'"
    
    print(f"Processing sheet {sheet_name},\t index {i},\t for dscp {dscp}")
    sheet = workbook[sheet_name]

    # Get the start and end times
    start = constants.args.start[i]
    end = constants.args.end[i]

    # Add pair to the dictionary
    constants.start_end_times[sheet_name] = (start, end)

    #get the flow and hop latency, for the given dscp, that includes all flows and switches
    AVG_flows_latency, STD_flows_latency, AVG_hop_latency, STD_hop_latency = get_avg_stdev_flow_hop_latency(start, end, dscp_condition)

    # % of packets that went to each individual switch (switch_id)
    switch_data = get_byte_sum(start, end, dscp, dscp_condition)
    switch_data = calculate_percentages(start, end, switch_data, dscp, dscp_condition)
    switch_data = get_mean_standard_deviation(switch_data, dscp)

    write_INT_results(sheet, AVG_flows_latency, STD_flows_latency, AVG_hop_latency, STD_hop_latency, dscp)
    write_INT_results_switchID(sheet, switch_data, dscp)


def set_caculation_section():
    # Configure each sheet
    workbook = load_workbook(constants.final_file_path)

    # For each sheet and respectice file, see the time interval given, get the values from the DB, and set the values in the sheet
    # Get nº each sheet
    for i, sheet_name in enumerate(workbook.sheetnames): 
        scemario = sheet_name.split("-")[0]
        scenario_DSCPs = constants.DSCP_per_scenario[scemario]
        for current_dscp in scenario_DSCPs:
            set_caculation_formulas(workbook, sheet_name, current_dscp, scenario_DSCPs)
            set_INT_results(workbook, sheet_name, current_dscp, i)               #technically, also contains another section, but its easier to call it here
    
    # Save the workbook
    workbook.save(constants.final_file_path)

def set_compare_non_Emergency_to_Emergency_variation():
    # Configure each sheet
    workbook = load_workbook(constants.final_file_path)

    for i, sheet in enumerate(workbook.sheetnames):
        #can i can not exceed the number of args.f (last one is comparasions)
        if i >= len(constants.args.f):
            break

        sheet = workbook[sheet]
        sheet.append([""])
        sheet.append([""])
        sheet.append([""])
        sheet.append([""])

        # Set new headers
        max_line = sheet.max_row
        sheet[f'A{max_line + 2}'] = "Flows Types"
        sheet[f'B{max_line + 2}'] = "Non-Emergency Flows"
        sheet[f'C{max_line + 2}'] = "Emergency Flows"
        sheet[f'D{max_line + 2}'] = "Variation (%)"
        sheet[f'E{max_line + 3}'] = "DSCP"
        
        sheet[f'A{max_line + 3}'] = "AVG 1º Packet Delay (nanoseconds)"
        sheet[f'A{max_line + 4}'] = "AVG Flow Delay (nanoseconds)"

        sheet[f'A{max_line + 2}'].font = Font(bold=True)
        sheet[f'B{max_line + 2}'].font = Font(bold=True)
        sheet[f'C{max_line + 2}'].font = Font(bold=True)
        sheet[f'D{max_line + 2}'].font = Font(bold=True)
        sheet[f'E{max_line + 2}'].font = Font(bold=True)

        sheet[f'A{max_line + 3}'].font = Font(bold=True)
        sheet[f'A{max_line + 4}'].font = Font(bold=True)

        start = constants.args.start[i]
        end = constants.args.end[i]

        # Define the row range of data to consider
        row_range = constants.last_line_raw_data[sheet.title]
        after_raw_data = row_range + 2

        # Set the formula for the Non-Emergency Flows
        sheet[f'B{max_line + 3}'] = f'=ROUND(AVERAGEIF(D1:D{row_range}, "<40" , O1:O{row_range}), 2'  
        sheet[f'B{max_line + 4}'] = f'=ROUND(AVERAGEIFS(B{after_raw_data}:B{max_line}, A{after_raw_data}:A{max_line}, "AVG Flows Latency (nanoseconds)", E{after_raw_data}:E{max_line}, "<40"), 2)'  #NOT IDEAL to do avg of avg, but inflix stores all tags as strings including dscp, making it difficult to apply this logic in a query

        # Set the formula for the Emergency Flows
        sheet[f'C{max_line + 3}'] = f'=ROUND(AVERAGEIF(D1:D{row_range}, ">=40", O1:O{row_range}), 2'
        sheet[f'C{max_line + 4}'] = f'=ROUND(AVERAGEIFS(B{after_raw_data}:B{max_line}, A{after_raw_data}:A{max_line}, "AVG Flows Latency (nanoseconds)", E{after_raw_data}:E{max_line}, ">=40"), 2)'

        #Set comparasion formulas, for the AVG 1º Packet Delay and AVG Flow Delay in percentage
        sheet[f'D{max_line + 3}'] = f'=IFERROR(ROUND((C{max_line + 3} - B{max_line + 3})/ABS(B{max_line + 3}) * 100, 2), "none")'
        sheet[f'D{max_line + 4}'] = f'=IFERROR(ROUND((C{max_line + 4} - B{max_line + 4})/ABS(B{max_line + 4}) * 100, 2), "none")'

        sheet[f'E{max_line + 3}'] = -1
        sheet[f'E{max_line + 4}'] = -1

    workbook.save(constants.final_file_path)


def configure_final_file():

    # Raw data area
    set_pkt_loss()
    set_fist_pkt_delay()
    
    # Calculations area for each dscp 
    set_caculation_section()

    set_compare_non_Emergency_to_Emergency_variation()
    comparasion_sheet.set_Comparison_sheet()            #Configure Comparasion sheet
    graphs.create_graphs()                              #create the graphs

